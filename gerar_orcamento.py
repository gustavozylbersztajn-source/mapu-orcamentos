#!/usr/bin/env python3
"""
MAPU — Gerador Automático de Orçamentos
Uso: python3 gerar_orcamento.py
"""

import os
import re
import io
import zipfile
import shutil
import json
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime
from pathlib import Path

try:
    from fpdf import FPDF
except ImportError:
    os.system("pip3 install fpdf2 --quiet")
    from fpdf import FPDF

import openpyxl

# ── PATHS ──────────────────────────────────────────────────────────────────────
MAPU_ROOT   = Path.home() / "Dropbox" / "Family Room" / "4.MAPU"
ORCAMENTOS  = MAPU_ROOT / "4.ORCAMENTOS"
PRECOS_XLSX  = MAPU_ROOT / "3.PLANEJAMENTO" / "1.PRECIFICAÇÃO" / "PRECOS_CABINS.xlsx"
PLANNER_XLSX = MAPU_ROOT / "3.PLANEJAMENTO" / "PLANEJAMENTO_GERAL" / "MAPU_PLANNER.xlsx"

# Assets bundled — usados quando caminhos locais não existem (nuvem)
_ASSETS      = Path(__file__).parent / "assets"
_ASSETS_FONTS = _ASSETS / "fonts"

LOGO_PATH   = MAPU_ROOT / "1.MKT" / "3.IDENTIDADEVISUAL" / "1.LOGOS" / "MAPU_logo_BADGEblacksml3.png"
LOGO_WHITE  = MAPU_ROOT / "1.MKT" / "3.IDENTIDADEVISUAL" / "1.LOGOS" / "MAPU_logo_BADGEwhitesml.png"
if not LOGO_PATH.exists():
    LOGO_PATH  = _ASSETS / "logos" / "MAPU_logo_BADGEblacksml3.png"
    LOGO_WHITE = _ASSETS / "logos" / "MAPU_logo_BADGEwhitesml.png"

_FONTS      = Path.home() / "Library" / "Fonts"
FONT_DISPLAY = _FONTS / "FF_DIN_Condensed_Black.otf"
FONT_HEAVY   = _FONTS / "FF_DIN_Condensed_Bold.otf"
FONT_MEDIUM  = _FONTS / "FF_DIN_Condensed_Regular.otf"
if not FONT_DISPLAY.exists():
    FONT_DISPLAY = _ASSETS_FONTS / "FF_DIN_Condensed_Black.otf"
    FONT_HEAVY   = _ASSETS_FONTS / "FF_DIN_Condensed_Bold.otf"
    FONT_MEDIUM  = _ASSETS_FONTS / "FF_DIN_Condensed_Regular.otf"

# Termos e Condições — anexados ao final de cada proposta, no idioma do orçamento
_POLICIES_DIR = MAPU_ROOT / "2.OPERACIONAL" / "1.POLICIES"
POLICY_PDF = {
    "pt": _POLICIES_DIR / "PORT" / "MAPU_LODGE_SPA_Termos_Condicoes_PT.pdf",
    "es": _POLICIES_DIR / "ESP"  / "MAPU_LODGE_SPA_Terminos_Condiciones_ES.pdf",
    "en": _POLICIES_DIR / "ING"  / "MAPU_LODGE_SPA_Terms_Conditions_EN.pdf",
}
if not POLICY_PDF["en"].exists():
    POLICY_PDF = {
        "pt": _ASSETS / "policies" / "MAPU_LODGE_SPA_Termos_Condicoes_PT.pdf",
        "es": _ASSETS / "policies" / "MAPU_LODGE_SPA_Terminos_Condiciones_ES.pdf",
        "en": _ASSETS / "policies" / "MAPU_LODGE_SPA_Terms_Conditions_EN.pdf",
    }

# Modo nuvem: sem acesso ao Dropbox local
IS_CLOUD = not MAPU_ROOT.exists()

# Prefixo usado nos paths do Dropbox API — o app agora tem acesso Full Dropbox
# (antes era App folder, escopado automaticamente a Apps/mapu-orcamentos-agencias)
DROPBOX_ROOT = "/Family Room/4.MAPU/4.ORCAMENTOS"

_BG    = (26, 26, 26)
_WHITE = (255, 255, 255)
_GL    = (180, 180, 180)
_GD    = (115, 115, 115)
_LINE  = (65, 65, 65)
_BOX   = (42, 42, 42)

SUBFOLDERS = [
    "1. ORCAMENTO", "2. DOCUMENTOS", "3. RESERVA",
    "4. CONSUMOS", "5. PAGAMENTOS", "6. BOLETAS", "7. INVOICE & RECEIPT",
]

MARKUP       = 1.2
CC_RATE      = 0.05
INFANT_SUPPL = 50_000

CONFIG_PATH = ORCAMENTOS / "config.json"

_DEFAULT_MEAL_PRICES = {
    "breakfast": {"adult": 24_000, "child": 16_800},
    "dinner":    {"adult": 55_000, "child": 38_000},
}

def _load_meal_prices():
    return _DEFAULT_MEAL_PRICES

CABIN_MAP = {
    "COIGUE": {
        (1, 0): ("C", "BUDGET COIGUE 1"),
        (2, 0): ("D", "BUDGET COIGUE 2"),
        (2, 1): ("E", "BUDGET COIGUE 2+1"),
    },
    "NIRE": {
        (1, 0): ("F", "BUDGET NIRE 1"),
        (2, 0): ("G", "BUDGET NIRE 2"),
        (2, 1): ("H", "BUDGET NIRE 2+1"),
        (2, 2): ("I", "BUDGET NIRE 2+2"),
    },
    "CHAITEN": {
        (1, 0): ("J", "BUDGET CHAITEN 1"),
        (2, 0): ("K", "BUDGET CHAITEN 2"),
        (3, 0): ("L", "BUDGET CHAITEN 3"),
        (2, 1): ("M", "BUDGET CHAITEN 2+1"),
        (2, 2): ("N", "BUDGET CHAITEN 2+2"),
    },
    "CORCOVADO": {
        (1, 0): ("O", "BUDGET CORCOVADO 1"),
        (2, 0): ("P", "BUDGET CORCOVADO 2"),
        (2, 1): ("Q", "BUDGET CORCOVADO 2+1"),
        (2, 2): ("R", "BUDGET CORCOVADO 2+2"),
        (2, 3): ("S", "BUDGET CORCOVADO 2+3"),
        (3, 0): ("T", "BUDGET CORCOVADO 3"),
        (3, 1): ("U", "BUDGET CORCOVADO 3+1"),
        (3, 2): ("V", "BUDGET CORCOVADO 3+2"),
        (4, 0): ("W", "BUDGET CORCOVADO 4"),
        (4, 1): ("X", "BUDGET CORCOVADO 4+1"),
        (5, 0): ("Y", "BUDGET CORCOVADO 5"),
    },
}

BASEPAX_ROWS = {
    "breakfast": 6,
    "lunch":     22,
    "dinner":    23,
}

STANDARD_OCCUPANCY = {
    "COIGUE":    2,
    "NIRE":      2,
    "CHAITEN":   2,
    "CORCOVADO": 4,
}

# ── CATÁLOGOS TIPO 2 — (nome, row BASE PAX) ────────────────────────────────────
ACTIVITIES_PER_USE = [
    ("Hottube",                37),
]

ACTIVITIES_PER_PAX = [
    ("Trekking 1/2 dia",       38),
    ("Trekking dia inteiro",   39),
    ("Flotada familiar",       40),
    ("Rafting 1/2 dia",        41),
    ("Rafting dia inteiro",    42),
    ("Cavalgada 1/2 dia",      43),
    ("Cavalgada dia inteiro",  44),
    ("Lancha Yelcho",          45),
    ("Recreação infantil 3hs", 46),
]

DRINKS_ITEMS = [
    ("Sucos",    27),
    ("Cervejas", 29),
    ("Vinhos",   30),
    ("Drinks",   31),
    ("Cooler",   32),
]

TRANSPORT_ITEMS = [
    ("Transfer chegada aeroporto", 50),
    ("Transfer partida aeroporto", 51),
    ("Transfer interno",           52),
    ("Uber",                       53),
]

EXPERIENCES_ITEMS = [
    ("WS Instructor", 58),
]

EXTRAS_ITEMS = [
    ("Coordenação produção",   64),
    ("Mão de obra extra",      65),
    ("Limpeza check-in/out",   66),
    ("Limpeza diária",         67),
    ("Souvenir",               68),
    ("Assistente coordenação", 69),
]

QUINCHO_ITEMS = [
    ("Quincho day use", 73),
]

# ── MAPEAMENTO CALENDAR → BASE PAX ────────────────────────────────────────────
# (regex, categoria, base_pax_row)
# categoria: 'breakfast'|'lunch'|'dinner'|'drinks'|'activity'|'hottube'|'transport'|'experience'|'extra'|'quincho'
CALENDAR_MAP = [
    (r'caf[eé]|breakfast|desayuno',                  "breakfast",   6),
    (r'almo[cç]o|lunch|almuerzo',                   "lunch",       7),
    (r'jantar|dinner|cordeiro|salm[aã]o|hamburgue|pizza|empanada|tasting|welcome', "dinner", None),
    (r'bebidas?|drinks?|vinho|wine|cerveja|beer|suco|juice|cooler', "drinks", 14),
    (r'cavalgada.*(½|1/2|meio|half)',                "activity",   38),
    (r'cavalgada.*(1 dia|full)',                     "activity",   39),
    (r'rafting.*(½|1/2|meio|half)',                  "activity",   40),
    (r'rafting.*(1 dia|full)',                       "activity",   41),
    (r'trekking.*(½|1/2|meio|half)',                 "activity",   42),
    (r'trekking.*(1 dia|full)',                      "activity",   43),
    (r'pesca',                                       "activity",   44),
    (r'flotada',                                     "activity",   45),
    (r'recrea[cç][aã]o|infantil',                    "activity",   46),
    (r'hot.?tub|hottube|ofur[uú]',                  "hottube",    37),
    (r'transfer.*(chegada|arrival)',                 "transport",  50),
    (r'transfer.*(partida|departure|saída)',         "transport",  51),
    (r'transfer.*intern|traslado.*intern',           "transport",  52),
    (r'uber',                                        "transport",  53),
    (r'ws|workshop|instructor',                      "experience", 58),
    (r'coordena[cç][aã]o.*produ[cç]|produ[cç].*coordena', "extra", 64),
    (r'quincho',                                     "quincho",   73),
]

# Jantar → rows sequenciais B15-B19
DINNER_ROWS = [15, 16, 17, 18, 19]

# ── TRADUÇÕES ──────────────────────────────────────────────────────────────────
LANG_MAP = {
    "português": "pt", "Português": "pt", "pt": "pt",
    "español":   "es", "Español":   "es", "es": "es",
    "english":   "en", "English":   "en", "en": "en",
}

TRANSLATIONS = {
    "pt": {
        "months":      {1:"JAN",2:"FEV",3:"MAR",4:"ABR",5:"MAI",6:"JUN",
                        7:"JUL",8:"AGO",9:"SET",10:"OUT",11:"NOV",12:"DEZ"},
        "titulo":      "PROPOSTA",
        "adultos":     "ADULTOS",
        "crianca":     "CRIANÇA",
        "criancas":    "CRIANÇAS",
        "noites":      "NOITES",
        "cabana":      "Cabana",
        "adultos_l":   "adultos",
        "crianca_l":   "criança",
        "criancas_l":  "crianças",
        "pessoas_l":      "pessoas",
        "passageiros_l":  "passageiros",
        "sec_hosp":    "HOSPEDAGEM",
        "desconto":    "Desconto",
        "sec_equipe":  "EQUIPE MAPU",
        "equipe_desc": "Equipe dedicada + operacional",
        "sec_food":    "ALIMENTAÇÃO",
        "sec_bebidas": "BEBIDAS",
        "sec_ativ":    "ATIVIDADES",
        "sec_transp":  "TRANSPORTE",
        "sec_exp":     "EXPERIÊNCIAS",
        "sec_extras":  "EXTRAS",
        "sec_quincho": "QUINCHO",
        "qty_vezes":   "vez" ,
        "qty_vezes_pl":"vezes",
        "qty_usos":    "uso",
        "qty_usos_pl": "usos",
        "sec_resumo":  "RESUMO",
        "sec_opc":     "OPCIONAIS — NÃO INCLUÍDOS",
        "res_hosp":    "Hospedagem",
        "res_hosp_mp": "Hospedagem + Refeições",
        "res_food":    "Alimentação",
        "res_bebidas": "Bebidas",
        "res_ativ":    "Atividades",
        "res_transp":  "Transportes + Traslados",
        "res_exp":     "Experiências",
        "res_extras":  "Extras",
        "res_quincho": "Quincho",
        "res_equipe":  "Equipe MAPU",
        "res_agencia":  "Taxa agência",
        "res_cc":      "Taxa cartão de crédito (CC)",
        "res_desconto": "Desconto {pct}%",
        "res_subtotal": "Subtotal neto",
        "res_iva":     "IVA 19% (se aplicável)",
        "total_label": "TOTAL  CLP",
        "total_label_moeda": "(Pesos Chilenos)",
        "total_cc_note": "*Valor com taxa de cartão de crédito (CC) incluída",
        "usd_ref":     "VALOR TOTAL EQUIVALENTE EM USD (câmbio {rate} CLP/USD)",
        "por_adulto":  "Por adulto (÷ {n})",
        "rodape":      "Valores em CLP · Válido 7 dias",
        "formas_pagto_title": "FORMAS DE PAGAMENTO",
        "formas_pagto_metodos": "Cartão de crédito (Visa/Mastercard/Amex) ou transferência bancária internacional.",
        "formas_pagto_prazo": "25% de sinal para confirmar a reserva ({sinal}). Saldo de 75% ({saldo}) até o check-in.",
        "food_note_breakfast": "Café da manhã (opcional): CL$ {adult} por adulto e CL$ {child} por criança, por noite — não incluído no valor total, cobrado à parte caso desejem adicionar.",
        "food_note_dinner": "Jantar: em breve teremos definidas as opções de jantar para alta temporada 2027 — assim que confirmado, compartilharemos os detalhes com vocês.",
        "pagamentos_title": "PAGAMENTOS",
        "saldo_label": "Saldo a pagar",
        "balance_due_note": "Saldo a pagar até {date}",
        "bank_title":   "DADOS PARA TRANSFERÊNCIA",
        "bank_bank":    "Banco",
        "bank_holder":  "Titular",
        "bank_account": "Conta corrente",
        "bank_rut":     "RUT",
        "bank_swift":   "SWIFT",
        "bank_address": "Endereço",
        "bank_email":   "Email",
        "meals": {
            "breakfast": "Café da manhã",
            "dinner":    "Jantar",
        },
        "bkf_unit":  "dia",
        "bkf_units": "dias",
        "meal_plans": {
            "half": "Meia pensão — café da manhã e jantar",
            "bkf":  "Apenas café da manhã",
            "none": "Sem alimentação",
        },
        "opcionais": [
            ("Canoa",              "$ 50.000 CLP / embarcação (até 4hs)"),
            ("Stand Up Paddle",    "$ 30.000 CLP / embarcação (até 2hs)"),
            ("Hot Tub com snacks", "$ 60.000 CLP / uso"),
        ],
        "prefix": "Proposta",
        "email_subject": "Sua Proposta MAPU [living in harmony] — {client}",
        "email_greeting": "Olá {client},",
        "email_intro": "Em anexo PDF para análise de sua proposta personalizada para uma estadia exclusiva no MAPU [living in harmony], aqui na Patagonia Chilena.",
        "email_details": "Detalhes da sua estadia:",
        "email_checkin": "Check-in",
        "email_checkout": "Check-out",
        "email_nights": "Noites",
        "email_cabin": "Cabana",
        "email_guests": "Hóspedes",
        "email_total": "TOTAL CLP",
        "email_validity": "Esta proposta é válida por 7 dias.",
        "email_cta": "Para confirmar sua reserva ou tirar dúvidas, responda este email ou entre em contato conosco.",
        "email_signature": "Equipe MAPU",
        "email_adults": "adultos",
        "email_children": "crianças",
        "email_open_pdf": "Veja sua proposta no PDF em anexo.",
    },
    "es": {
        "months":      {1:"ENE",2:"FEB",3:"MAR",4:"ABR",5:"MAY",6:"JUN",
                        7:"JUL",8:"AGO",9:"SEP",10:"OCT",11:"NOV",12:"DIC"},
        "titulo":      "PROPUESTA",
        "adultos":     "ADULTOS",
        "crianca":     "NIÑO",
        "criancas":    "NIÑOS",
        "noites":      "NOCHES",
        "cabana":      "Cabaña",
        "adultos_l":   "adultos",
        "crianca_l":   "niño",
        "criancas_l":  "niños",
        "pessoas_l":      "personas",
        "passageiros_l":  "pasajeros",
        "sec_hosp":    "ALOJAMIENTO",
        "desconto":    "Descuento",
        "sec_equipe":  "EQUIPO MAPU",
        "equipe_desc": "Equipo dedicado + operacional",
        "sec_food":    "ALIMENTACIÓN",
        "sec_bebidas": "BEBIDAS",
        "sec_ativ":    "ACTIVIDADES",
        "sec_transp":  "TRANSPORTE",
        "sec_exp":     "EXPERIENCIAS",
        "sec_extras":  "EXTRAS",
        "sec_quincho": "QUINCHO",
        "qty_vezes":   "vez",
        "qty_vezes_pl":"veces",
        "qty_usos":    "uso",
        "qty_usos_pl": "usos",
        "sec_resumo":  "RESUMEN",
        "sec_opc":     "OPCIONALES — NO INCLUIDOS",
        "res_hosp":    "Alojamiento",
        "res_hosp_mp": "Alojamiento + Comidas",
        "res_food":    "Alimentación",
        "res_bebidas": "Bebidas",
        "res_ativ":    "Actividades",
        "res_transp":  "Transportes + Traslados",
        "res_exp":     "Experiencias",
        "res_extras":  "Extras",
        "res_quincho": "Quincho",
        "res_equipe":  "Equipo MAPU",
        "res_agencia": "Comisión agencia",
        "res_desconto": "Descuento {pct}%",
        "res_cc":      "Tarifa tarjeta de crédito (CC)",
        "res_subtotal":"Subtotal neto",
        "res_iva":     "IVA 19% (se aplicável)",
        "total_label": "TOTAL  CLP",
        "total_label_moeda": "(Pesos Chilenos)",
        "total_cc_note": "*Valor con tarifa de tarjeta de crédito (CC) incluida",
        "usd_ref":     "VALOR TOTAL EQUIVALENTE EN USD (cambio {rate} CLP/USD)",
        "por_adulto":  "Por adulto (÷ {n})",
        "rodape":      "Valores en CLP · Válido 7 días",
        "formas_pagto_title": "FORMAS DE PAGO",
        "formas_pagto_metodos": "Tarjeta de crédito (Visa/Mastercard/Amex) o transferencia bancaria internacional.",
        "formas_pagto_prazo": "25% de anticipo para confirmar la reserva ({sinal}). Saldo de 75% ({saldo}) hasta el check-in.",
        "food_note_breakfast": "Desayuno (opcional): CL$ {adult} por adulto y CL$ {child} por niño, por noche — no incluido en el valor total, se cobra aparte si desean agregarlo.",
        "food_note_dinner": "Cena: pronto tendremos definidas las opciones de cena para la temporada alta 2027 — en cuanto esté confirmado, compartiremos los detalles con ustedes.",
        "pagamentos_title": "PAGOS",
        "saldo_label": "Saldo a pagar",
        "balance_due_note": "Saldo a pagar hasta el {date}",
        "bank_title":   "DATOS PARA TRANSFERENCIA",
        "bank_bank":    "Banco",
        "bank_holder":  "Titular",
        "bank_account": "Cuenta corriente",
        "bank_rut":     "RUT",
        "bank_swift":   "SWIFT",
        "bank_address": "Dirección",
        "bank_email":   "Email",
        "meals": {
            "breakfast": "Desayuno",
            "dinner":    "Cena",
        },
        "bkf_unit":  "día",
        "bkf_units": "días",
        "meal_plans": {
            "half": "Media pensión — desayuno y cena",
            "bkf":  "Solo desayuno",
            "none": "Sin alimentación",
        },
        "opcionais": [
            ("Canoa",              "$ 50.000 CLP / embarcación (hasta 4hs)"),
            ("Stand Up Paddle",    "$ 30.000 CLP / embarcación (hasta 2hs)"),
            ("Hot Tub con snacks", "$ 60.000 CLP / uso"),
        ],
        "prefix": "Propuesta",
        "email_subject": "Tu Propuesta MAPU [living in harmony] — {client}",
        "email_greeting": "Hola {client},",
        "email_intro": "Adjuntamos en PDF tu propuesta personalizada para una estadía exclusiva en MAPU [living in harmony], aquí en la Patagonia Chilena.",
        "email_details": "Detalles de tu estadía:",
        "email_checkin": "Check-in",
        "email_checkout": "Check-out",
        "email_nights": "Noches",
        "email_cabin": "Cabaña",
        "email_guests": "Huéspedes",
        "email_total": "TOTAL CLP",
        "email_validity": "Esta propuesta es válida por 7 días.",
        "email_cta": "Para confirmar tu reserva o resolver dudas, responde este email o contáctanos.",
        "email_signature": "Equipo MAPU",
        "email_adults": "adultos",
        "email_children": "niños",
        "email_open_pdf": "Vea su propuesta en el PDF adjunto.",
    },
    "en": {
        "months":      {1:"JAN",2:"FEB",3:"MAR",4:"APR",5:"MAY",6:"JUN",
                        7:"JUL",8:"AUG",9:"SEP",10:"OCT",11:"NOV",12:"DEC"},
        "titulo":      "PROPOSAL",
        "adultos":     "ADULTS",
        "crianca":     "CHILD",
        "criancas":    "CHILDREN",
        "noites":      "NIGHTS",
        "cabana":      "Cabin",
        "adultos_l":   "adults",
        "crianca_l":   "child",
        "criancas_l":  "children",
        "pessoas_l":      "people",
        "passageiros_l":  "passengers",
        "sec_hosp":    "ACCOMMODATION",
        "desconto":    "Discount",
        "sec_equipe":  "MAPU TEAM",
        "equipe_desc": "Dedicated team + operations",
        "sec_food":    "MEALS",
        "sec_bebidas": "BEVERAGES",
        "sec_ativ":    "ACTIVITIES",
        "sec_transp":  "TRANSPORT",
        "sec_exp":     "EXPERIENCES",
        "sec_extras":  "EXTRAS",
        "sec_quincho": "QUINCHO",
        "qty_vezes":   "time",
        "qty_vezes_pl":"times",
        "qty_usos":    "use",
        "qty_usos_pl": "uses",
        "sec_resumo":  "SUMMARY",
        "sec_opc":     "OPTIONAL — NOT INCLUDED",
        "res_hosp":    "Accommodation",
        "res_hosp_mp": "Accommodation + Meals",
        "res_food":    "Meals",
        "res_bebidas": "Beverages",
        "res_ativ":    "Activities",
        "res_transp":  "Transport + Transfers",
        "res_exp":     "Experiences",
        "res_extras":  "Extras",
        "res_quincho": "Quincho",
        "res_equipe":  "MAPU Team",
        "res_agencia": "Agency fee",
        "res_desconto": "Discount {pct}%",
        "res_cc":      "Credit card fee (CC)",
        "res_subtotal":"Net subtotal",
        "res_iva":     "VAT 19% (if applicable)",
        "total_label": "TOTAL  CLP",
        "total_label_moeda": "(Chilean Pesos)",
        "total_cc_note": "*Amount includes credit card fee (CC)",
        "usd_ref":     "TOTAL USD EQUIVALENT VALUE (rate {rate} CLP/USD)",
        "por_adulto":  "Per adult (÷ {n})",
        "rodape":      "Values in CLP · Valid 7 days",
        "formas_pagto_title": "PAYMENT METHODS",
        "formas_pagto_metodos": "Credit card (Visa/Mastercard/Amex) or international bank transfer.",
        "formas_pagto_prazo": "25% deposit to confirm the reservation ({sinal}). Remaining 75% balance ({saldo}) due by check-in.",
        "food_note_breakfast": "Breakfast (optional): CL$ {adult} per adult and CL$ {child} per child, per night — not included in the total amount, charged separately if you'd like to add it.",
        "food_note_dinner": "Dinner: we'll soon have the dinner options defined for the 2027 high season — as soon as confirmed, we'll share the details with you.",
        "pagamentos_title": "PAYMENTS",
        "saldo_label": "Balance due",
        "balance_due_note": "Balance due by {date}",
        "bank_title":   "BANK TRANSFER DETAILS",
        "bank_bank":    "Bank",
        "bank_holder":  "Account holder",
        "bank_account": "Checking account",
        "bank_rut":     "Tax ID (RUT)",
        "bank_swift":   "SWIFT",
        "bank_address": "Address",
        "bank_email":   "Email",
        "meals": {
            "breakfast": "Breakfast",
            "dinner":    "Dinner",
        },
        "bkf_unit":  "day",
        "bkf_units": "days",
        "meal_plans": {
            "half": "Half board — breakfast & dinner",
            "bkf":  "Breakfast only",
            "none": "No meals included",
        },
        "opcionais": [
            ("Canoe",               "$ 50,000 CLP / boat (up to 4hs)"),
            ("Stand Up Paddle",     "$ 30,000 CLP / board (up to 2hs)"),
            ("Hot Tub with snacks", "$ 60,000 CLP / use"),
        ],
        "prefix": "Proposal",
        "email_subject": "Your MAPU [living in harmony] Proposal — {client}",
        "email_greeting": "Hello {client},",
        "email_intro": "Please find attached the PDF with your personalized proposal for an exclusive stay at MAPU [living in harmony], here in Chilean Patagonia.",
        "email_details": "Your stay details:",
        "email_checkin": "Check-in",
        "email_checkout": "Check-out",
        "email_nights": "Nights",
        "email_cabin": "Cabin",
        "email_guests": "Guests",
        "email_total": "TOTAL CLP",
        "email_validity": "This proposal is valid for 7 days.",
        "email_cta": "To confirm your booking or for any questions, simply reply to this email or contact us.",
        "email_signature": "MAPU Team",
        "email_adults": "adults",
        "email_children": "children",
        "email_open_pdf": "View your proposal in the attached PDF.",
    },
}

SMTP_CONFIG_PATH = Path.home() / ".mapu_config.json"


# ── HELPERS ────────────────────────────────────────────────────────────────────
def fmt_clp(v):
    return "$ " + f"{int(round(v)):,}".replace(",", ".")

def fmt_usd(v):
    return "$ " + f"{int(round(v)):,}".replace(",", ".")

def date_fmt(d, lang="pt"):
    t = TRANSLATIONS.get(lang, TRANSLATIONS["pt"])
    return f"{d.day:02d} {t['months'][d.month]} {d.year}"

def safe(text):
    return (text
        .replace("—", "-").replace("–", "-")
        .replace("→", ">").replace("←", "<")
        .replace("ó", "o").replace("ú", "u").replace("á", "a")
        .replace("é", "e").replace("í", "i").replace("ñ", "n")
        .replace("ã", "a").replace("ç", "c").replace("ô", "o")
        .replace("Á", "A").replace("É", "E").replace("Í", "I")
        .replace("Ó", "O").replace("Ú", "U").replace("Ñ", "N")
        .replace("ü", "u").replace("Ü", "U")
    )

def _select_items(catalog, header, qty_label=None):
    """Mostra lista numerada, retorna dict {row: info} para itens selecionados.
    Se qty_label=None, selecionar = B=1 (sem pergunta de quantidade).
    """
    print(f"\n── {header} {'─'*(48-len(header))}")
    for i, (name, _row) in enumerate(catalog, 1):
        print(f"  {i:>2}. {name}")
    sel = prompt("  Selecionar (ex: 1,3 | 0 = nenhum)", "0").strip()
    if sel == "0" or not sel:
        return {}
    result = {}
    try:
        indices = [int(x.strip()) for x in sel.replace(" ", ",").split(",") if x.strip()]
    except ValueError:
        return {}
    for idx in indices:
        if 1 <= idx <= len(catalog):
            name, row = catalog[idx - 1]
            if qty_label:
                qty = prompt_int(f"    {name} — {qty_label}", 1)
            else:
                qty = 1
            if qty > 0:
                result[row] = {"name": name, "qty": qty}
    return result


def prompt(msg, default=None):
    suf = f" (default: {default})" if default is not None else ""
    v = input(f"  {msg}{suf}: ").strip()
    return v if v else (str(default) if default is not None else "")

def prompt_int(msg, default=0):
    try:
        return int(prompt(msg, default))
    except ValueError:
        return default

def _find_template_dropbox():
    """Baixa o template canônico (4.ORCAMENTOS/2027/template/planilha_orcamentos_python.xlsx)
    direto do Dropbox pra um arquivo temporário. Mantém uma única planilha base real —
    o app na nuvem não depende mais de uma cópia bundled que pode ficar desatualizada."""
    dbx = _get_dropbox_client()
    if not dbx:
        return None
    try:
        import tempfile
        _, response = dbx.files_download(
            f"{DROPBOX_ROOT}/2027/template/planilha_orcamentos_python.xlsx")
        tmp = Path(tempfile.gettempdir()) / "mapu_template_dropbox.xlsx"
        tmp.write_bytes(response.content)
        return tmp
    except Exception as e:
        print(f"_find_template_dropbox erro: {e}")
        return None


def find_template(checkin_year):
    # Nuvem: tenta baixar o template ao vivo do Dropbox; cai pro bundled se falhar
    bundled = _ASSETS / "templates" / "planilha_orcamentos_python.xlsx"
    if IS_CLOUD:
        live = _find_template_dropbox()
        if live is not None:
            return live
        return bundled if bundled.exists() else None
    search_years = sorted(set([checkin_year, checkin_year - 1, 2027]), reverse=True)
    for y in search_years:
        for name in ["planilha_orcamentos_python.xlsx", "ANOMESDIA_nomecliente_B.xlsx", "ANOMESDIA_nomecliente.xlsx"]:
            p = ORCAMENTOS / str(y) / "template" / name
            if p.exists():
                return p
    p = ORCAMENTOS / "FOLDERS" / "1.ORCAMENTO" / "ANOMESDIA_nomecliente.xlsx"
    if p.exists():
        return p
    return bundled if bundled.exists() else None


def _fetch_exchange_rate(default=900):
    apis = [
        ("https://open.er-api.com/v6/latest/USD",      lambda d: round(d["rates"]["CLP"])),
        ("https://api.frankfurter.app/latest?from=USD&to=CLP", lambda d: round(d["rates"]["CLP"])),
    ]
    import urllib.request, json
    for url, extract in apis:
        try:
            with urllib.request.urlopen(url, timeout=5) as r:
                return extract(json.loads(r.read()))
        except Exception:
            continue
    print(f"  ⚠ Câmbio não disponível — usando {default}")
    return default


# ── CARREGAR PREÇOS ────────────────────────────────────────────────────────────
def _load_precos_json(is_alta):
    """Último fallback: lê preços de precos.json quando PLANNER não está disponível nem local nem via Dropbox."""
    for p in (Path(__file__).parent / "precos.json", _ASSETS / "precos.json"):
        if p.exists():
            data = json.loads(p.read_text())
            key = "alta" if is_alta else "baixa"
            return data[key]
    raise FileNotFoundError("precos.json não encontrado")


def _extract_precos_from_workbook(wb, is_alta):
    """Lê os preços da aba ⚙️ INPUTS de um workbook MAPU_PLANNER já aberto (local ou baixado do Dropbox).
    Preço da cabana fica SEM IVA — o IVA só é aplicado depois, em calculate(), e só quando
    chilean_client=True (turista estrangeiro é isento). Nunca embutir IVA aqui."""
    ws = wb["⚙️ INPUTS"]
    iva_rate     = ws["B32"].value or 0.19
    cc_rate      = ws["B33"].value or 0.04
    agency_rate  = ws["B34"].value or 0.20

    cabin_order = ["COIGUE", "NIRE", "CHAITEN", "CORCOVADO"]
    base = {}
    for i, cabin in enumerate(cabin_order, start=6):
        sem_iva = (ws[f"B{i}"].value if is_alta else ws[f"C{i}"].value) or 0
        base[cabin] = round(sem_iva)

    prices = {
        "standard":    base,
        "long":        {k: round(v * 0.95) for k, v in base.items()},
        "agency_rate": agency_rate,
        "cc_rate":     cc_rate,
        "iva_rate":    iva_rate,
    }

    bkf_adult = (ws["B16"].value if is_alta else ws["C16"].value) or 24_000
    bkf_child = (ws["B17"].value if is_alta else ws["C17"].value) or 16_800
    din_adult = ws["B21"].value or 55_000
    din_child = ws["B22"].value or 38_000
    prices["meal_prices"] = {
        "breakfast": {"adult": bkf_adult, "child": bkf_child},
        "dinner":    {"adult": din_adult, "child": din_child},
    }
    return prices


def _load_precos_dropbox(is_alta):
    """Baixa MAPU_PLANNER.xlsx do Dropbox via API e lê os preços — usado quando não há acesso
    ao filesystem local (nuvem). Mantém o app sincronizado com a planilha viva sem precisar
    regenerar precos.json manualmente a cada alteração de preço."""
    dbx = _get_dropbox_client()
    if not dbx:
        return None
    try:
        import io
        _, response = dbx.files_download(
            "/Family Room/4.MAPU/3.PLANEJAMENTO/PLANEJAMENTO_GERAL/MAPU_PLANNER.xlsx")
        wb = openpyxl.load_workbook(io.BytesIO(response.content), data_only=True)
        return _extract_precos_from_workbook(wb, is_alta)
    except Exception as e:
        print(f"_load_precos_dropbox erro: {e}")
        return None


def load_precos(checkin_year=None, checkin_month=None):
    """Lê preços de MAPU_PLANNER.xlsx — Alta (Dez·Jan·Fev·Mar) vs Baixa (Abr–Nov).
    Ordem de prioridade: arquivo local → Dropbox API (nuvem) → precos.json (último recurso)."""
    is_alta = checkin_month in (1, 2, 3, 12) if checkin_month else False

    if not IS_CLOUD and PLANNER_XLSX.exists():
        wb = openpyxl.load_workbook(str(PLANNER_XLSX), data_only=True)
        return _extract_precos_from_workbook(wb, is_alta)

    dropbox_prices = _load_precos_dropbox(is_alta)
    if dropbox_prices is not None:
        return dropbox_prices

    return _load_precos_json(is_alta)


# ── COLETAR INPUTS ─────────────────────────────────────────────────────────────
def collect_inputs():
    print("\n" + "━"*54)
    print("   MAPU — GERADOR DE ORÇAMENTO")
    print("━"*54 + "\n")

    client_raw  = prompt("Nome do cliente (ex: MariaEmilia)").strip()
    client_slug = client_raw.upper().replace(" ", "")

    checkin_s  = prompt("Check-in  (DD/MM/AA)")
    checkout_s = prompt("Check-out (DD/MM/AA)")
    def _parse_date(s):
        s = s.strip()
        if len(s) == 6 and s.isdigit():
            s = f"{s[:2]}/{s[2:4]}/{s[4:]}"
        elif len(s) == 8 and s.isdigit():
            s = f"{s[:2]}/{s[2:4]}/{s[4:]}"
        for fmt in ("%d/%m/%Y", "%d/%m/%y"):
            try: return datetime.strptime(s, fmt).date()
            except ValueError: pass
        raise ValueError(f"Data inválida: {s}")
    checkin  = _parse_date(checkin_s)
    checkout = _parse_date(checkout_s)
    nights   = (checkout - checkin).days
    if nights < 3:
        print(f"\n  ⚠  Mínimo 3 noites — período informado tem {nights}n. Confirme as datas.")
    print(f"\n  → {nights} noites\n")

    print("── CABANAS ─────────────────────────────────────────")
    cabins = {}
    for cabin in ["COIGUE", "NIRE", "CHAITEN", "CORCOVADO"]:
        adults = prompt_int(f"{cabin} — adultos (0 = não usar)")
        if adults > 0:
            infants = prompt_int(f"{cabin} — crianças/bebês")
            cabins[cabin] = {"adults": adults, "infants": infants}

    total_adults  = sum(c["adults"]  for c in cabins.values())
    total_infants = sum(c["infants"] for c in cabins.values())
    print(f"\n  → Total: {total_adults} adultos + {total_infants} crianças\n")

    exchange = _fetch_exchange_rate()

    print("\n── TIPO DE ORÇAMENTO ────────────────────────────────")
    print("    1 = Básico   (hospedagem + alimentação)")
    print("    2 = Completo (+ bebidas, atividades, transporte, experiências, extras, quincho)")
    _tipo = prompt("  Tipo", "1").strip()
    tipo  = 2 if _tipo == "2" else 1

    print("\n── REFEIÇÕES ────────────────────────────────────────")
    meals = {}
    # Café da manhã é opcional, por pessoa. Jantar ainda não é oferecido (serviço de comida a definir a partir de meio de novembro).
    _inc_cafe = prompt("  Incluir café da manhã? (s/n)", "s").strip().lower()
    if _inc_cafe in ("s", "sim", "y", "yes"):
        print(f"  Café da manhã × {nights} noites")
        meals["breakfast"] = nights
    else:
        print("  Sem café da manhã")

    extras_sel = {}
    if tipo == 2:
        _act = {}
        _act.update(_select_items(ACTIVITIES_PER_USE, "ATIVIDADES — HOTTUBE", "quantidade de usos"))
        _act.update(_select_items(ACTIVITIES_PER_PAX, "ATIVIDADES — DEMAIS"))
        per_pax_rows = {row for _, row in ACTIVITIES_PER_PAX}
        for row in per_pax_rows:
            if row in _act:
                _act[row]["is_per_pax"] = True
        extras_sel["activities"] = _act
        extras_sel["drinks"]     = _select_items(
            DRINKS_ITEMS, "BEBIDAS", "quantidade por cabana")
        _transp = _select_items(TRANSPORT_ITEMS, "TRANSPORTE")
        for info in _transp.values():
            info["is_transfer"] = True
        extras_sel["transport"] = _transp
        extras_sel["experiences"]= _select_items(
            EXPERIENCES_ITEMS, "EXPERIÊNCIAS")
        extras_sel["extras"]     = _select_items(
            EXTRAS_ITEMS, "EXTRAS")
        extras_sel["quincho"]    = _select_items(
            QUINCHO_ITEMS, "QUINCHO", "usos")

    mapu_team = 0

    agency         = prompt("\nCom agência? (s/n)", "n").lower() == "s"
    chilean_client = False  # IVA sempre mostrado como informativo, nunca obrigatório

    adjustments = {}
    if cabins and prompt("Ajuste de tarifa? (s/n)", "n").lower() == "s":
        print("  → % por cabana  (ex: 25 = +25% sobretarifa | -15 = -15% desconto):")
        for cabin in cabins:
            pct = prompt(f"    {cabin} %", "0").strip()
            try:
                val = float(pct)
                if val != 0:
                    adjustments[cabin] = val / 100
            except ValueError:
                pass

    print("\n── IDIOMA ───────────────────────────────────────────")
    _lang_opt = prompt("Idioma: 1=pt / 2=es / 3=en", "1").strip()
    lang = {"1": "pt", "2": "es", "3": "en"}.get(_lang_opt, LANG_MAP.get(_lang_opt, "pt"))

    print("\n── CONTATO ──────────────────────────────────────────")
    email_cliente    = prompt("Email do cliente (deixe vazio para não enviar)", "").strip()
    telefone_cliente = prompt("Telefone do cliente (deixe vazio para omitir)", "").strip()

    return {
        "client_raw":        client_raw,
        "client":            client_slug,
        "checkin":           checkin,
        "checkout":          checkout,
        "nights":            nights,
        "cabins":            cabins,
        "total_adults":      total_adults,
        "total_infants":     total_infants,
        "exchange":          exchange,
        "meals":             meals,
        "mapu_team":         mapu_team,
        "agency":            agency,
        "chilean_client":    chilean_client,
        "lang":              lang,
        "adjustments":       adjustments,
        "email_cliente":     email_cliente,
        "telefone_cliente":  telefone_cliente,
        "tipo":              tipo,
        "extras_sel":        extras_sel,
    }


# ── CALCULAR PREÇOS ────────────────────────────────────────────────────────────
def calculate(data, prices):
    nights   = data["nights"]
    exchange = data["exchange"]
    adults   = data["total_adults"]
    infants  = data.get("total_infants", 0)
    meals    = data["meals"]
    lang     = data.get("lang", "pt")
    CC_RATE  = prices.get("cc_rate", 0.04)
    t        = TRANSLATIONS.get(lang, TRANSLATIONS["pt"])

    tier    = "long" if nights >= 7 else "standard"
    lodging = 0
    discount_total  = 0
    cabin_breakdown = {}
    for cabin, pax in data["cabins"].items():
        clp        = prices[tier][cabin]
        adj        = data.get("adjustments", {}).get(cabin, 0)
        total_pax  = pax["adults"] + pax["infants"]
        extra_supl = INFANT_SUPPL if total_pax > STANDARD_OCCUPANCY.get(cabin, 2) else 0
        base_amt   = (clp + extra_supl) * nights
        adj_amt    = round(base_amt * adj)
        cabin_amt  = base_amt + adj_amt
        lodging   += cabin_amt
        cabin_discount = 0
        if adj < 0:
            cabin_discount  = round(abs(adj_amt))
            discount_total += cabin_discount
        cabin_breakdown[cabin] = {"adults": pax["adults"], "infants": pax["infants"], "amount": round(cabin_amt), "discount": cabin_discount}

    food           = 0
    food_lines     = []
    food_breakdown = []

    mp = prices.get("meal_prices") or _load_meal_prices()

    if meals.get("breakfast", 0) > 0:
        qty     = meals["breakfast"]
        bkf_adult = mp["breakfast"]["adult"] * adults * qty
        bkf_child = mp["breakfast"]["child"] * infants * qty
        bkf_amt   = bkf_adult + bkf_child
        food   += bkf_amt
        food_lines.append(f"{t['meals']['breakfast']} × {qty}")
        food_breakdown.append({
            "label":       t["meals"]["breakfast"],
            "qty":         qty,
            "amount":      round(bkf_amt),
            "adult_total": round(bkf_adult),
            "child_total": round(bkf_child),
        })

    if meals.get("dinner", 0) > 0:
        qty     = meals["dinner"]
        din_adult = mp["dinner"]["adult"] * adults * qty
        din_child = mp["dinner"]["child"] * infants * qty
        din_amt   = din_adult + din_child
        food   += din_amt
        food_lines.append(f"{t['meals']['dinner']} × {qty}")
        food_breakdown.append({
            "label":       t["meals"]["dinner"],
            "qty":         qty,
            "amount":      round(din_amt),
            "adult_total": round(din_adult),
            "child_total": round(din_child),
        })

    n_cabins = len(data["cabins"])

    has_b = meals.get("breakfast", 0) > 0
    has_d = meals.get("dinner", 0) > 0
    if has_b and has_d:
        meal_plan_key = "half"
    elif has_b:
        meal_plan_key = "bkf"
    else:
        meal_plan_key = "none"
    meal_plan = t["meal_plans"][meal_plan_key]

    # Hospedagem e Alimentação ficam separadas no PDF (item próprio cada um) — Jul/2026
    food_embedded = False
    if food > 0:
        for cabin, pax in data["cabins"].items():
            c_adults  = pax["adults"]
            c_infants = pax["infants"]
            c_food = 0
            if meals.get("breakfast", 0) > 0:
                qty = meals["breakfast"]
                c_food += mp["breakfast"]["adult"] * c_adults  * qty
                c_food += mp["breakfast"]["child"]  * c_infants * qty
            if meals.get("dinner", 0) > 0:
                qty = meals["dinner"]
                c_food += mp["dinner"]["adult"] * c_adults  * qty
                c_food += mp["dinner"]["child"]  * c_infants * qty
            cabin_breakdown[cabin]["food_amount"] = round(c_food)
            cabin_breakdown[cabin]["meal_plan"]   = meal_plan

    mapu_team   = data["mapu_team"]
    subtotal_pre_discount = lodging + food + mapu_team
    extra_discount_pct = data.get("extra_discount_pct", 0) or 0
    extra_discount_amt = round(subtotal_pre_discount * extra_discount_pct)
    total_neto  = subtotal_pre_discount - extra_discount_amt
    iva         = round(total_neto * prices.get("iva_rate", 0.19)) if data.get("chilean_client") else 0
    total_bruto = total_neto + iva
    _agency_rate = data.get("agency_rate_override")
    if _agency_rate is None:
        _agency_rate = prices.get("agency_rate", 0.20)
    agency_fee  = round(total_bruto * _agency_rate) if data.get("agency") else 0
    total_cc    = (total_bruto + agency_fee) * (1 + CC_RATE)
    cc_fee      = round(total_cc - (total_bruto + agency_fee))
    usd_ref     = total_cc / exchange
    per_adult   = total_cc / adults if adults > 0 else 0

    return {
        "lodging":          lodging,
        "discount_total":   discount_total,
        "cabin_breakdown":  cabin_breakdown,
        "food":             food,
        "food_lines":       food_lines,
        "food_breakdown":   food_breakdown,
        "meal_plan":        meal_plan,
        "food_embedded":    food_embedded,
        "meal_prices":      mp,
        "mapu_team":      mapu_team,
        "extra_discount_pct": extra_discount_pct,
        "extra_discount_amt": extra_discount_amt,
        "total":          total_neto,
        "iva":            iva,
        "total_bruto":    total_bruto,
        "agency_fee":     agency_fee,
        "cc_fee":         cc_fee,
        "cc_rate":        CC_RATE,
        "total_cc":       total_cc,
        "usd_ref":        usd_ref,
        "per_adult":      per_adult,
    }


# ── CRIAR ESTRUTURA DE PASTAS ──────────────────────────────────────────────────
def create_folders(data):
    import tempfile
    slug = data["checkin"].strftime("%Y%m%d") + "_" + data["client"]
    if IS_CLOUD:
        path = Path(tempfile.gettempdir()) / "mapu_orcamentos" / slug
    else:
        year  = data["checkin"].strftime("%Y")
        month = data["checkin"].strftime("%m")
        path  = ORCAMENTOS / year / month / slug
    for sub in SUBFOLDERS:
        (path / sub).mkdir(parents=True, exist_ok=True)
    return path, slug


# ── POPULAR EXCEL ──────────────────────────────────────────────────────────────
def _build_sheet_map(template_path):
    import zipfile, xml.etree.ElementTree as ET
    NS_WB = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    NS_R  = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    with zipfile.ZipFile(template_path, "r") as z:
        wb_xml   = z.read("xl/workbook.xml")
        rels_xml = z.read("xl/_rels/workbook.xml.rels")
    rid_to_file = {
        r.get("Id"): (r.get("Target").lstrip("/") if r.get("Target", "").startswith("/") else "xl/" + r.get("Target"))
        for r in ET.fromstring(rels_xml)
    }
    sheet_map = {}
    for s in ET.fromstring(wb_xml).find(f"{{{NS_WB}}}sheets"):
        name = s.get("name")
        rid  = s.get(f"{{{NS_R}}}id")
        if name and rid and rid in rid_to_file:
            sheet_map[name] = rid_to_file[rid]
    return sheet_map


def _hide_inactive_budget_cols(xml_str, active_cols):
    import re
    active_nums = {ord(c) - ord("A") + 1 for c in active_cols}

    def _expand_col(m):
        tag = m.group(0)
        mn = re.search(r'min="(\d+)"', tag)
        mx = re.search(r'max="(\d+)"', tag)
        if not mn or not mx:
            return tag
        col_min, col_max = int(mn.group(1)), int(mx.group(1))
        # só atua nas colunas de cabana (C=3 a Y=25)
        if col_min < 3 or col_max > 25:
            return tag
        # se range multi-coluna, expande em cols individuais para ocultar seletivamente
        base = re.sub(r'\s+hidden="[^"]*"', "", tag)
        result = []
        for i in range(col_min, col_max + 1):
            col_tag = re.sub(r'min="\d+"', f'min="{i}"', base)
            col_tag = re.sub(r'max="\d+"', f'max="{i}"', col_tag)
            if i not in active_nums:
                col_tag = col_tag.replace("/>", ' hidden="1"/>')
            result.append(col_tag)
        return "".join(result)

    return re.sub(r'<col\b[^/]*/>', _expand_col, xml_str)


def _clear_range(xml_str, col_start, col_end, row_start, row_end):
    """Remove todas as células no retângulo [col_start..col_end] x [row_start..row_end]."""
    import re
    refs = set()
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            col_letter = ""
            n = c
            while n > 0:
                n, rem = divmod(n - 1, 26)
                col_letter = chr(65 + rem) + col_letter
            refs.add(f"{col_letter}{r}")
    ATTR = r'(?:\s+[\w:]+="[^"]*")'
    def remove_cell(m):
        ref = re.search(r'r="([^"]+)"', m.group(0))
        return "" if (ref and ref.group(1) in refs) else m.group(0)
    return re.sub(
        rf'<c\s+r="[^"]+"{ATTR}*\s*(?:/>|>.*?</c>)',
        remove_cell,
        xml_str,
        flags=re.DOTALL,
    )


def _set_cell(xml_str, cell_ref, value):
    import re
    from datetime import date as date_type

    ATTR    = r'(?:\s+[\w:]+="[^"]*")'
    pattern = re.compile(
        rf'<c\s+r="{re.escape(cell_ref)}"{ATTR}*\s*(?:/>|>.*?</c>)',
        re.DOTALL
    )
    m = pattern.search(xml_str)
    if not m:
        return xml_str

    old       = m.group(0)
    open_m    = re.match(rf'<c((?:\s+[\w:]+="[^"]*")*)\s*', old)
    raw_attrs = open_m.group(1) if open_m else ""
    keep      = re.sub(r'\s+r="[^"]*"', "", raw_attrs)
    keep      = re.sub(r'\s+t="[^"]*"', "", keep)

    if isinstance(value, date_type):
        serial   = (value - date_type(1899, 12, 30)).days
        new_cell = f'<c r="{cell_ref}"{keep}><v>{serial}</v></c>'
    elif isinstance(value, str) and value.startswith("="):
        # Fórmula — escreve sem cached value para forçar recálculo
        formula = value[1:].replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
        new_cell = f'<c r="{cell_ref}"{keep}><f>{formula}</f><v></v></c>'
    elif isinstance(value, str):
        escaped  = value.replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
        new_cell = f'<c r="{cell_ref}"{keep} t="inlineStr"><is><t>{escaped}</t></is></c>'
    else:
        iv       = int(value) if float(value) == int(value) else value
        new_cell = f'<c r="{cell_ref}"{keep}><v>{iv}</v></c>'

    return xml_str[: m.start()] + new_cell + xml_str[m.end() :]


def _reorder_budget_tabs(xml_str, active_sheet_names):
    """Move as abas BUDGET das cabanas selecionadas pra logo à direita de RESULTADO."""
    import re
    if not active_sheet_names:
        return xml_str
    m = re.search(r'(<sheets>)(.*?)(</sheets>)', xml_str, re.DOTALL)
    if not m:
        return xml_str
    tags = re.findall(r'<sheet\b[^>]*/>', m.group(2))

    def _name(tag):
        nm = re.search(r'name="([^"]*)"', tag)
        return nm.group(1) if nm else None

    rest_tags   = [t for t in tags if _name(t) not in active_sheet_names]
    active_tags = [t for t in tags if _name(t) in active_sheet_names]
    active_tags.sort(key=lambda t: active_sheet_names.index(_name(t)))

    result_idx = next((i for i, t in enumerate(rest_tags) if _name(t) == "RESULTADO"), None)
    if result_idx is None:
        return xml_str

    new_order = rest_tags[: result_idx + 1] + active_tags + rest_tags[result_idx + 1 :]
    new_body  = "".join(new_order)
    return xml_str[: m.start()] + m.group(1) + new_body + m.group(3) + xml_str[m.end() :]


def populate_excel(data, client_path, slug, calcs, prices=None):
    if prices is None:
        prices = load_precos(checkin_year=data["checkin"].year, checkin_month=data["checkin"].month)
    import zipfile, os as _os

    template = find_template(data["checkin"].year)
    if not template:
        print("  ⚠ Template XLSX não encontrado — pulando Excel.")
        return None

    out = client_path / "1. ORCAMENTO" / f"{slug}.xlsx"
    shutil.copy2(str(template), str(out))

    sheet_map = _build_sheet_map(str(template))
    nights    = data["nights"]
    meals     = data["meals"]
    updates   = {}
    clears    = {}   # sheet_path → [(col_start, col_end, row_start, row_end)]

    if "CALENDAR" in sheet_map:
        cal_cells = {
            "C2": data["checkin"],
            "C3": data["checkout"],
            "L2": data["client_raw"],   # J2 tem fórmula C3-C2 (noites) — não sobreescrever
        }
        updates[sheet_map["CALENDAR"]] = cal_cells
        # Limpa bloco D7:P23 (programa anterior) — D=4, P=16
        clears[sheet_map["CALENDAR"]] = [(4, 16, 7, 23)]

    active_cols = []
    if "BUDGET RESUMO" in sheet_map:
        flags = {f"{col}8": 0 for col in "CDEFGHIJKLMNOPQRSTUVWXY"}
        for cabin, pax in data["cabins"].items():
            key = (pax["adults"], pax["infants"])
            cm  = CABIN_MAP.get(cabin, {})
            if key not in cm:
                key = (pax["adults"], 0)
            if key in cm:
                col, _ = cm[key]
                flags[f"{col}8"] = 1
                active_cols.append(col)
        # C23: comissão agência — 0 para reserva direta, taxa do PLANNER (ou override do admin) se com agência
        _agency_rate = data.get("agency_rate_override")
        if _agency_rate is None:
            _agency_rate = prices.get("agency_rate", 0.20)
        flags["C23"] = _agency_rate if data.get("agency") else 0
        flags["C24"] = prices.get("cc_rate", 0.04)
        updates[sheet_map["BUDGET RESUMO"]] = flags

    if "BASE PAX" in sheet_map:
        bp_cells = {}
        for food_key, bp_row in BASEPAX_ROWS.items():
            if food_key == "breakfast":
                bp_cells[f"B{bp_row}"] = 1 if meals.get(food_key, 0) > 0 else 0
            else:
                bp_cells[f"B{bp_row}"] = meals.get(food_key, 0)
        bp_cells["B30"] = 0

        # Preços de refeição da temporada → BASE PAX G6/H6 (café) e G23/H23 (jantar)
        mp = calcs.get("meal_prices", _DEFAULT_MEAL_PRICES)
        bp_cells["G6"]  = mp["breakfast"]["adult"]
        bp_cells["H6"]  = mp["breakfast"]["child"]
        bp_cells["G23"] = mp["dinner"]["adult"]
        bp_cells["H23"] = mp["dinner"]["child"]
        for row in (64, 65, 66, 67):
            bp_cells[f"C{row}"] = 0

        # Tipo 2 — escreve quantidades das categorias extras
        total_pax = data.get("total_adults", 1) + data.get("total_infants", 0)
        hottube_row = ACTIVITIES_PER_USE[0][1]  # row 37
        for category, items in data.get("extras_sel", {}).items():
            for row, info in items.items():
                qty = info["qty"]
                if row == hottube_row:
                    # Hottube é flat fee — divide por pax para neutralizar multiplicação do BUDGET
                    bp_cells[f"B{row}"] = round(qty / max(total_pax, 1), 6)
                else:
                    bp_cells[f"B{row}"] = qty

        updates[sheet_map["BASE PAX"]] = bp_cells

    active_sheet_names = []
    for cabin, pax in data["cabins"].items():
        key = (pax["adults"], pax["infants"])
        cm  = CABIN_MAP.get(cabin, {})
        if key not in cm:
            key = (pax["adults"], 0)
        if key not in cm:
            continue
        _, sheet_name = cm[key]
        if sheet_name not in sheet_map:
            continue
        active_sheet_names.append(sheet_name)
        total_pax  = pax["adults"] + pax["infants"]
        needs_extra = total_pax > STANDARD_OCCUPANCY.get(cabin, 2)
        std_price  = prices["standard"][cabin]
        long_price = prices["long"][cabin]
        cc_rate    = prices.get("cc_rate", 0.04)
        cells   = {
            "B5": pax["adults"], "B6": pax["infants"],
            "B15": 0, "B16": 0, "B17": 0, "B18": 0, "B19": 0, "B20": 0,
            "B24": 0, "B25": 0, "B26": 0, "B27": 0,
            # C12 calibrado: RESULTADO aplica fator 0.95 → C12 = 1/0.95-1 faz D×0.95 = B exato
            "C12": round(1/0.95 - 1, 8),
            # Zera adição indevida de comissão no jantar
            "E47": 0, "E48": 0,
            # Corrige fórmulas café/jantar para incluir crianças (template 1-adulto tem fórmula incompleta)
            "I31": "='BASE PAX'!B6*B9*($B$5*'BASE PAX'!G6+$B$6*'BASE PAX'!H6)",
            "I48": "='BASE PAX'!B23*($B$5*'BASE PAX'!G23+$B$6*'BASE PAX'!H23)",
            "C24": std_price,  "C25": std_price,
            "C26": long_price, "C27": long_price,
            "D25": INFANT_SUPPL, "D27": INFANT_SUPPL,
            "F24": 0, "F25": 0, "F26": 0, "F27": 0,
        }
        cells["B25" if needs_extra else "B24"] = 1 if nights <= 6 else 0
        cells["B27" if needs_extra else "B26"] = 1 if nights > 6  else 0
        adj = data.get("adjustments", {}).get(cabin, 0)
        if adj:
            f_val = -adj  # Excel: G = E*(1-F) → F negativo = sobretarifa, F positivo = desconto
            cells["F24"] = f_val
            cells["F25"] = f_val
            cells["F26"] = f_val
            cells["F27"] = f_val
        updates[sheet_map[sheet_name]] = cells

    br_file = sheet_map.get("BUDGET RESUMO", "")
    tmp = str(out) + ".tmp"
    with zipfile.ZipFile(str(out), "r") as zin:
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                raw = zin.read(item.filename)
                if item.filename in updates or item.filename in clears:
                    xml = raw.decode("utf-8")
                    for col_s, col_e, row_s, row_e in clears.get(item.filename, []):
                        xml = _clear_range(xml, col_s, col_e, row_s, row_e)
                    for cell_ref, val in updates.get(item.filename, {}).items():
                        xml = _set_cell(xml, cell_ref, val)
                    if item.filename == br_file and active_cols:
                        xml = _hide_inactive_budget_cols(xml, active_cols)
                    raw = xml.encode("utf-8")
                elif item.filename == "xl/calcChain.xml":
                    continue  # remove stale calcChain — Excel regenera ao abrir
                elif item.filename == "[Content_Types].xml":
                    import re as _re
                    xml = raw.decode("utf-8")
                    xml = _re.sub(r'<Override[^>]*calcChain[^>]*/>', '', xml)
                    raw = xml.encode("utf-8")
                elif item.filename == "xl/_rels/workbook.xml.rels":
                    import re as _re
                    xml = raw.decode("utf-8")
                    xml = _re.sub(r'<Relationship[^>]*calcChain[^>]*/>', '', xml)
                    raw = xml.encode("utf-8")
                elif item.filename == "xl/workbook.xml":
                    import re as _re
                    xml = raw.decode("utf-8")
                    if "fullCalcOnLoad" not in xml:
                        xml = _re.sub(r'<calcPr\b([^/]*)/>',
                                      r'<calcPr\1 fullCalcOnLoad="1"/>',
                                      xml)
                    xml = _reorder_budget_tabs(xml, active_sheet_names)
                    raw = xml.encode("utf-8")
                zout.writestr(item, raw)
    _os.replace(tmp, str(out))
    print(f"  ✓ Excel salvo: {out.name}")
    return out


# ── GERAR PDF ──────────────────────────────────────────────────────────────────
def generate_pdf(data, calcs, client_path, slug, prop_num):
    lang = data.get("lang", "pt")
    t    = TRANSLATIONS.get(lang, TRANSLATIONS["pt"])

    out = (client_path / "1. ORCAMENTO"
           / f"{t['prefix']}_{data['client']}_{data['checkin'].strftime('%Y%m%d')}.pdf")

    nights   = data["nights"]
    adults   = data["total_adults"]
    infants  = data["total_infants"]
    checkin  = data["checkin"]
    checkout = data["checkout"]
    exchange = data["exchange"]
    W, H     = 297.0, 210.0

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_margins(0, 0, 0)
    pdf.set_auto_page_break(False)
    pdf.add_page()
    pdf.add_font("Disp",   "", str(FONT_DISPLAY))
    pdf.add_font("Heavy",  "", str(FONT_HEAVY))
    pdf.add_font("Medium", "", str(FONT_MEDIUM))

    pdf.set_fill_color(*_BG)
    pdf.rect(-1, -1, W + 2, H + 2, "F")

    # HEADER
    from datetime import date as _date
    today = _date.today()
    pdf.set_font("Heavy", size=9)
    pdf.set_text_color(*_GL)
    pdf.set_xy(0, 11)
    pdf.cell(W, 4, f"Nº {checkin.strftime('%Y%m%d')}-{prop_num}  ·  {today.strftime('%d/%m/%Y')}", align="C")

    pdf.set_font("Disp", size=27)
    pdf.set_text_color(*_WHITE)
    pdf.set_xy(0, 17)
    pdf.cell(W, 12, f"{t['titulo']} · {data['client_raw'].upper()}", align="C")

    if infants:
        pax_str = (f"{adults} {t['adultos']} + "
                   f"{infants} {t['crianca'] if infants == 1 else t['criancas']}")
    else:
        pax_str = f"{adults} {t['adultos']}"

    sub = (f"CHECK-IN {date_fmt(checkin, lang)}  ·  "
           f"CHECK-OUT {date_fmt(checkout, lang)}  ·  "
           f"{nights} {t['noites']}  ·  {pax_str}")
    pdf.set_font("Heavy", size=8)
    pdf.set_text_color(*_GL)
    pdf.set_xy(0, 32)
    pdf.cell(W, 4, sub, align="C")

    email    = data.get("email_cliente", "")
    telefone = data.get("telefone_cliente", "")
    if email or telefone:
        pdf.set_font("Heavy", size=8.5)
        pdf.set_text_color(*_GL)
        pdf.set_xy(0, 37)
        if email and telefone:
            sep   = "  ·  "
            wa_num = "".join(c for c in telefone if c.isdigit())
            wa_url = f"https://wa.me/{wa_num}"
            w_email = pdf.get_string_width(email)
            w_sep   = pdf.get_string_width(sep)
            w_tel   = pdf.get_string_width(telefone)
            total_w = w_email + w_sep + w_tel
            x_start = (W - total_w) / 2
            pdf.set_xy(x_start, 37)
            pdf.cell(w_email, 4, email)
            pdf.cell(w_sep,   4, sep)
            pdf.cell(w_tel,   4, telefone, link=wa_url)
        elif email:
            pdf.cell(W, 4, email, align="C")
        else:
            wa_num = "".join(c for c in telefone if c.isdigit())
            wa_url = f"https://wa.me/{wa_num}"
            w_tel  = pdf.get_string_width(telefone)
            pdf.set_xy((W - w_tel) / 2, 37)
            pdf.cell(w_tel, 4, telefone, link=wa_url)
        _divider_y = 55
        _content_y = 57.0
    else:
        _divider_y = 39
        _content_y = 41.0


    tipo = data.get("tipo", 1)
    if tipo == 2:
        # 3 colunas: esquerda | meio | resumo
        LX, LW = 10.0, 91.0
        MX, MW = LX + LW + 8.0, 91.0
        RX      = MX + MW + 8.0
        RW      = W - RX - 10.0
        ly = my = ry = _content_y
    else:
        LX, LW = 20.0, 140.0
        RX      = LX + LW + 10.0
        RW      = W - RX - 20.0
        MX, MW = RX, RW  # não usado no Tipo 1
        ly = ry = _content_y
        my = _content_y

    def sec(x, y, w, label, subtitle=None):
        pdf.set_font("Heavy", size=7.5)
        pdf.set_text_color(*_WHITE)
        if subtitle:
            pdf.set_xy(x, y)
            pdf.cell(w * 0.38, 4.5, label)
            pdf.set_font("Medium", size=7)
            pdf.set_text_color(*_GL)
            pdf.set_xy(x + w * 0.38, y)
            pdf.cell(w * 0.62, 4.5, subtitle, align="R")
        else:
            pdf.set_xy(x, y)
            pdf.cell(w, 4.5, label)
        pdf.set_draw_color(*_LINE)
        pdf.line(x, y + 5.5, x + w, y + 5.5)
        return y + 9.0

    # HOSPEDAGEM (subtitle mostra plano de refeição quando embutido)
    _hosp_sub = calcs["meal_plan"] if calcs.get("food_embedded") else None
    ly = sec(LX, ly, LW, t["sec_hosp"], _hosp_sub)
    pdf.set_font("Medium", size=8)
    for cabin, info in calcs.get("cabin_breakdown", {}).items():
        a, i = info["adults"], info["infants"]
        if i:
            pax_desc = (f"{a:02d} {t['adultos_l']} + "
                        f"{i:02d} {t['crianca_l'] if i == 1 else t['criancas_l']}")
        else:
            pax_desc = f"{a:02d} {t['adultos_l']}"
        disc = info.get("discount", 0)
        pdf.set_text_color(*_GL)
        pdf.set_xy(LX, ly)
        pdf.cell(LW * 0.37, 4.5, f"{t['cabana']} {cabin.capitalize()}")
        pdf.set_xy(LX + LW * 0.37, ly)
        pdf.cell(LW * 0.25, 4.5, pax_desc)
        if disc > 0:
            adj_pct = data.get("adjustments", {}).get(cabin, 0)
            pct_str = f"{abs(adj_pct)*100:.0f}%"
            pdf.set_xy(LX + LW * 0.62, ly)
            pdf.cell(LW * 0.22, 4.5, f"(desc. {pct_str} · {fmt_clp(disc)})", align="R")
        pdf.set_xy(LX + LW * 0.84, ly)
        pdf.cell(LW * 0.16, 4.5, fmt_clp(info["amount"]), align="R")
        ly += 4.5
    ly += 2
    pdf.set_font("Heavy", size=11)
    pdf.set_text_color(*_WHITE)
    pdf.set_xy(LX, ly)
    pdf.cell(LW, 5.5, fmt_clp(calcs["lodging"]))
    ly += 10

    # EQUIPE MAPU
    if calcs["mapu_team"] > 0:
        ly = sec(LX, ly, LW, t["sec_equipe"])
        pdf.set_font("Medium", size=8)
        pdf.set_text_color(*_GL)
        pdf.set_xy(LX, ly)
        pdf.cell(LW * 0.65, 4.5, t["equipe_desc"])
        pdf.set_xy(LX + LW * 0.65, ly)
        pdf.cell(LW * 0.35, 4.5, fmt_clp(calcs["mapu_team"]), align="R")
        ly += 10

    # ALIMENTAÇÃO (só renderiza quando não embutido na hospedagem)
    if calcs.get("food", 0) > 0:
        ly = sec(LX, ly, LW, t["sec_food"], calcs["meal_plan"])
        pdf.set_font("Medium", size=8)
        pdf.set_text_color(*_GL)
        for fb in calcs.get("food_breakdown", []):
            pax_parts = []
            adults = data.get("total_adults", 0)
            infants = data.get("total_infants", 0)
            if adults:
                pax_parts.append(f"{adults} {t['adultos_l']}")
            if infants:
                lbl = t['crianca_l'] if infants == 1 else t['criancas_l']
                pax_parts.append(f"{infants} {lbl}")
            pax_str = f" ({' · '.join(pax_parts)})" if pax_parts else ""
            pdf.set_text_color(*_GL)
            pdf.set_xy(LX, ly)
            pdf.cell(LW * 0.37, 4.5, fb["label"])
            pdf.set_xy(LX + LW * 0.37, ly)
            pdf.cell(LW * 0.33, 4.5, f"× {fb['qty']}{pax_str}")
            pdf.set_xy(LX + LW * 0.70, ly)
            pdf.cell(LW * 0.30, 4.5, fmt_clp(fb["amount"]), align="R")
            ly += 4.5
        ly += 2
        pdf.set_font("Heavy", size=11)
        pdf.set_text_color(*_WHITE)
        pdf.set_xy(LX, ly)
        pdf.cell(LW, 5.5, fmt_clp(calcs["food"]))
        ly += 10
    elif data.get("food_notes"):
        # Sem refeição embutida no total, mas com informações/observações (ex: café opcional à parte, jantar a definir)
        ly = sec(LX, ly, LW, t["sec_food"], calcs["meal_plan"])
        pdf.set_font("Medium", size=7.5)
        pdf.set_text_color(*_GL)
        for note in data["food_notes"]:
            pdf.set_xy(LX, ly)
            pdf.multi_cell(LW, 4, note)
            ly = pdf.get_y() + 1
        ly += 6

    # SEÇÕES EXTRAS (Tipo 2)
    total_pax = data.get("total_adults", 0) + data.get("total_infants", 0)

    def _extras_section(sec_key, total_key, items_dict, qty_unit_s, qty_unit_pl, cx, cw, cy, show_qty=True):
        total = calcs.get(total_key, 0)
        if not items_dict and total <= 0:
            return cy
        cy = sec(cx, cy, cw, t[sec_key])
        pdf.set_font("Medium", size=8)
        for _row, info in items_dict.items():
            qty = info["qty"]
            if show_qty:
                if info.get("is_transfer"):
                    qty_str = f"× {total_pax} {t.get('passageiros_l', 'passageiros')}"
                elif info.get("is_per_pax"):
                    qty_str = f"× {total_pax} {t.get('pessoas_l', 'pessoas')}"
                else:
                    unit = qty_unit_s if qty == 1 else qty_unit_pl
                    qty_str = f"× {qty} {unit}"
            else:
                qty_str = ""
            pdf.set_text_color(*_GL)
            pdf.set_xy(cx, cy)
            pdf.cell(cw * 0.70, 4.5, info["name"])
            pdf.set_xy(cx + cw * 0.70, cy)
            pdf.cell(cw * 0.30, 4.5, qty_str, align="R")
            cy += 4.5
        if total > 0:
            cy += 2
            pdf.set_font("Heavy", size=11)
            pdf.set_text_color(*_WHITE)
            pdf.set_xy(cx, cy)
            pdf.cell(cw, 5.5, fmt_clp(total))
            cy += 10
        return cy

    extras_sel = data.get("extras_sel", {})
    vz, vzpl = t.get("qty_vezes", "vez"), t.get("qty_vezes_pl", "vezes")
    us, uspl   = t.get("qty_usos",  "uso"),          t.get("qty_usos_pl",  "usos")
    ucab, _    = "unid./cabana",                     "unid./cabana"

    if tipo == 2:
        # Col esquerda: bebidas + transporte (já tem hospedagem e alimentação acima)
        ly = _extras_section("sec_bebidas", "beverages",   extras_sel.get("drinks",      {}), ucab, ucab, LX, LW, ly)
        ly = _extras_section("sec_transp",  "transport",   extras_sel.get("transport",   {}), vz, vzpl, LX, LW, ly)
        # Col meio: atividades + experiências + extras
        my = _extras_section("sec_ativ",    "activities",  extras_sel.get("activities",  {}), vz, vzpl, MX, MW, my)
        my = _extras_section("sec_exp",     "experiences", extras_sel.get("experiences", {}), vz, vzpl, MX, MW, my)
        my = _extras_section("sec_extras",  "extras",      extras_sel.get("extras",      {}), us, uspl, MX, MW, my, show_qty=False)
        # Col direita: quincho antes do resumo
        ry = _extras_section("sec_quincho", "quincho",     extras_sel.get("quincho",     {}), us, uspl, RX, RW, ry, show_qty=False)
    else:
        ly = _extras_section("sec_ativ",    "activities",  extras_sel.get("activities",  {}), vz, vzpl, LX, LW, ly)
        ry = _extras_section("sec_bebidas", "beverages",   extras_sel.get("drinks",      {}), ucab, ucab, RX, RW, ry)
        ry = _extras_section("sec_transp",  "transport",   extras_sel.get("transport",   {}), vz, vzpl, RX, RW, ry)
        ry = _extras_section("sec_exp",     "experiences", extras_sel.get("experiences", {}), vz, vzpl, RX, RW, ry)
        ry = _extras_section("sec_extras",  "extras",      extras_sel.get("extras",      {}), us, uspl, RX, RW, ry, show_qty=False)
        ry = _extras_section("sec_quincho", "quincho",     extras_sel.get("quincho",     {}), us, uspl, RX, RW, ry, show_qty=False)

    # RESUMO
    ry = sec(RX, ry, RW, t["sec_resumo"])
    pdf.set_font("Medium", size=8)
    _hosp_label = t.get("res_hosp_mp", t["res_hosp"]) if calcs.get("food_embedded") else t["res_hosp"]
    resumo_lines = [(_hosp_label, calcs["lodging"])]
    if calcs.get("food",        0) > 0:
        resumo_lines.append((t["res_food"],    calcs["food"]))
    if calcs.get("beverages",   0) > 0:
        resumo_lines.append((t["res_bebidas"], calcs["beverages"]))
    if calcs.get("activities",  0) > 0:
        resumo_lines.append((t["res_ativ"],    calcs["activities"]))
    if calcs.get("transport",   0) > 0:
        resumo_lines.append((t["res_transp"],  calcs["transport"]))
    if calcs.get("experiences", 0) > 0:
        resumo_lines.append((t["res_exp"],     calcs["experiences"]))
    if calcs.get("extras",      0) > 0:
        resumo_lines.append((t["res_extras"],  calcs["extras"]))
    if calcs.get("quincho",     0) > 0:
        resumo_lines.append((t.get("res_quincho", "Quincho"), calcs["quincho"]))
    if calcs.get("extra_discount_amt", 0) > 0:
        pct_str = f"{calcs['extra_discount_pct']*100:.0f}"
        resumo_lines.append((t["res_desconto"].format(pct=pct_str), -calcs["extra_discount_amt"]))
    if calcs.get("agency_fee",  0) > 0:
        resumo_lines.append((t["res_agencia"], calcs["agency_fee"]))
    if calcs.get("cc_fee",      0) > 0:
        resumo_lines.append((t["res_cc"], calcs["cc_fee"]))
    for label, val in resumo_lines:
        pdf.set_text_color(*_GL)
        pdf.set_xy(RX, ry)
        pdf.cell(RW * 0.58, 4.5, label)
        pdf.set_xy(RX + RW * 0.58, ry)
        pdf.cell(RW * 0.42, 4.5, fmt_clp(val), align="R")
        ry += 4.5
    ry += 3

    hide_footer_info = data.get("hide_footer_info", False)

    # Caixa TOTAL (primeiro)
    box_h = 11.0
    pdf.set_fill_color(*_BOX)
    pdf.rect(RX - 2, ry, RW + 2, box_h, "F")
    pdf.set_font("Heavy", size=8)
    pdf.set_text_color(*_WHITE)
    pdf.set_xy(RX, ry + 1.5)
    pdf.cell(38, 6, t["total_label"])
    pdf.set_font("Medium", size=5.5)
    pdf.set_xy(RX, ry + 6.3)
    pdf.cell(38, 3, t["total_label_moeda"])
    pdf.set_font("Disp", size=17)
    pdf.set_xy(RX, ry + 0.5)
    pdf.cell(RW, box_h - 1, fmt_clp(calcs["total_cc"]), align="R")
    ry += box_h + 2.5

    if not hide_footer_info:
        # USD equivalente (logo depois do total)
        pdf.set_font("Heavy", size=6.8)
        pdf.set_text_color(*_GL)
        usd_val = "$ " + f"{int(round(calcs['usd_ref'])):,}".replace(",", ".") + " USD"
        pdf.set_xy(RX, ry)
        pdf.cell(RW * 0.58, 4.5, t["usd_ref"].format(rate=exchange))
        pdf.set_font("Heavy", size=8)
        pdf.set_xy(RX + RW * 0.58, ry)
        pdf.cell(RW * 0.42, 4.5, usd_val, align="R")
        ry += 6

        # IVA informativo
        iva_info = round(calcs["total"] * 0.19)
        pdf.set_font("Medium", size=8)
        pdf.set_text_color(*_WHITE)
        pdf.set_xy(RX, ry)
        pdf.cell(RW * 0.58, 4.5, t["res_iva"])
        pdf.set_xy(RX + RW * 0.58, ry)
        pdf.cell(RW * 0.42, 4.5, fmt_clp(iva_info), align="R")
        ry += 4.5

        ry += 4

    # FORMAS DE PAGAMENTO (padrão — aparece em todo orçamento) — por último
    if not hide_footer_info:
        pdf.set_font("Heavy", size=8)
        pdf.set_text_color(*_WHITE)
        pdf.set_xy(RX, ry)
        pdf.cell(RW, 4.5, t["formas_pagto_title"])
        ry += 5
        pdf.set_font("Medium", size=7)
        pdf.set_text_color(*_GL)
        pdf.set_xy(RX, ry)
        pdf.multi_cell(RW, 3.3, t["formas_pagto_metodos"])
        ry = pdf.get_y() + 0.5
        pdf.set_xy(RX, ry)
        sinal_val = round(calcs["total_cc"] * 0.25)
        saldo_val = calcs["total_cc"] - sinal_val
        pdf.multi_cell(RW, 3.3, t["formas_pagto_prazo"].format(
            sinal=fmt_clp(sinal_val), saldo=fmt_clp(saldo_val)))
        ry = pdf.get_y()
        ry += 3

    # OBSERVAÇÕES (opcional) — notas informativas extras (ex: café opcional não incluso, jantar a definir)
    notes = data.get("notes", [])
    if notes:
        ry += 4
        pdf.set_font("Medium", size=7)
        pdf.set_text_color(*_GL)
        for note in notes:
            pdf.set_xy(RX, ry)
            pdf.multi_cell(RW, 3.3, note, align="R")
            ry = pdf.get_y() + 1

    # PAGAMENTOS (opcional)
    payments = data.get("payments", [])
    if payments:
        ry += 8
        pdf.set_font("Heavy", size=8)
        pdf.set_text_color(*_WHITE)
        pdf.set_xy(RX, ry)
        pdf.cell(RW, 5, t["pagamentos_title"])
        pdf.set_draw_color(*_LINE)
        pdf.line(RX, ry + 6, RX + RW, ry + 6)
        ry += 10.0
        total_paid = 0
        for p in payments:
            pdf.set_font("Medium", size=9.5)
            pdf.set_text_color(*_GL)
            pdf.set_xy(RX, ry)
            pdf.cell(RW * 0.58, 5.5, p["label"])
            pdf.set_xy(RX + RW * 0.58, ry)
            pdf.cell(RW * 0.42, 5.5, fmt_clp(p["amount"]), align="R")
            total_paid += p["amount"]
            ry += 5.5
        saldo = calcs["total_cc"] - total_paid
        ry += 3
        pdf.set_font("Heavy", size=11)
        pdf.set_text_color(*_WHITE)
        pdf.set_xy(RX, ry)
        pdf.cell(RW * 0.58, 6, t["saldo_label"])
        pdf.set_xy(RX + RW * 0.58, ry)
        pdf.cell(RW * 0.42, 6, fmt_clp(saldo), align="R")
        ry += 8

        if data.get("balance_due_date"):
            pdf.set_font("Medium", size=7.5)
            pdf.set_text_color(*_GL)
            pdf.set_xy(RX, ry)
            pdf.cell(RW, 4, t["balance_due_note"].format(date=data["balance_due_date"]), align="R")
            ry += 6

        bank = data.get("bank_details")
        if bank:
            pdf.set_font("Heavy", size=7.5)
            pdf.set_text_color(*_WHITE)
            pdf.set_xy(RX, ry)
            pdf.cell(RW, 4, t["bank_title"], align="R")
            ry += 4.5
            pdf.set_font("Medium", size=7)
            pdf.set_text_color(*_GL)
            bank_fields = [
                (t["bank_bank"],    bank.get("bank")),
                (t["bank_holder"],  bank.get("holder")),
                (t["bank_account"], bank.get("account")),
                (t["bank_rut"],     bank.get("rut")),
                (t["bank_swift"],   bank.get("swift")),
                (t["bank_address"], bank.get("address")),
                (t["bank_email"],   bank.get("email")),
            ]
            for label, val in bank_fields:
                if not val:
                    continue
                pdf.set_xy(RX, ry)
                pdf.cell(RW, 3.8, f"{label}: {val}", align="R")
                ry += 3.8

    # FOOTER
    logo = LOGO_WHITE if LOGO_WHITE.exists() else LOGO_PATH
    if logo.exists():
        lw = 17
        pdf.image(str(logo), x=(W - lw) / 2, y=H - 26, w=lw)
    pdf.set_font("Heavy", size=9)
    pdf.set_text_color(*_GL)
    pdf.set_xy(LX, H - 18)
    pdf.cell(60, 5, "hola@mapuchile.com")
    pdf.set_xy(LX, H - 13)
    pdf.cell(60, 5, "+569 58642354")
    pdf.set_xy(W - 20 - 40, H - 13)
    pdf.cell(40, 5, "mapuchile.com", align="R")

    # Dados da agência no rodapé (lado direito)
    ag_name    = data.get("agency_name", "")
    ag_contact = data.get("agency_contact", "")
    ag_email   = data.get("agency_email", "")
    ag_phone   = data.get("agency_phone", "")
    if ag_name:
        pdf.set_font("Heavy", size=7.5)
        pdf.set_text_color(*_GL)
        line1 = f"{ag_name}  ·  {ag_contact}" if ag_contact else ag_name
        line2_parts = [p for p in (ag_email, ag_phone) if p]
        line2 = "  ·  ".join(line2_parts)
        pdf.set_xy(W - 20 - 100, H - 21)
        pdf.cell(100, 4, line1, align="R")
        if line2:
            pdf.set_xy(W - 20 - 100, H - 17)
            pdf.cell(100, 4, line2, align="R")

    pdf.output(str(out))

    if data.get("attach_policies", True):
        _append_policy_pdf(out, lang)

    print(f"  ✓ PDF salvo:  {out.name}")
    return out


def _append_policy_pdf(out_path, lang):
    """Anexa o PDF de Termos e Condições (no idioma do orçamento) ao final da proposta."""
    policy_path = POLICY_PDF.get(lang)
    if not policy_path or not policy_path.exists():
        return
    try:
        from pypdf import PdfReader, PdfWriter
    except ImportError:
        os.system("pip3 install pypdf --quiet")
        from pypdf import PdfReader, PdfWriter

    writer = PdfWriter()
    for p in (out_path, policy_path):
        reader = PdfReader(str(p))
        for page in reader.pages:
            writer.add_page(page)
    with open(out_path, "wb") as f:
        writer.write(f)


# ── EMAIL ─────────────────────────────────────────────────────────────────────
def load_smtp_config():
    if SMTP_CONFIG_PATH.exists():
        with open(SMTP_CONFIG_PATH) as f:
            return json.load(f)
    try:
        import streamlit as st
        s = st.secrets["smtp"]
        return {
            "smtp_host": s["smtp_host"],
            "smtp_port": int(s["smtp_port"]),
            "smtp_user": s["smtp_user"],
            "smtp_pass": s["smtp_pass"],
        }
    except Exception:
        return None


LOGO_EMAIL_URL = "https://www.dropbox.com/scl/fi/4oab735jgmbk1oiogrokk/MAPU_logo_BADGEwhitesml.png?rlkey=3ucdb7ncquwudwagk42wb8jmp&raw=1"


def _build_email_html(data, calcs, t):
    nights  = data["nights"]
    adults  = data["total_adults"]
    infants = data["total_infants"]
    client  = data["client_raw"]
    lang    = data.get("lang", "pt")

    if infants:
        guests_str = (f"{adults} {t['email_adults']} + "
                      f"{infants} {t['email_children']}")
    else:
        guests_str = f"{adults} {t['email_adults']}"

    cabins_str = ", ".join(
        f"{cab.capitalize()}" for cab in data["cabins"]
    )

    total_str = fmt_clp(calcs["total_cc"])

    return f"""<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f0f0f0;font-family:Arial,sans-serif">
<div style="max-width:580px;margin:32px auto;background:#fff">

  <!-- HEADER -->
  <div style="background:#000000;padding:32px;text-align:center">
    <img src="{LOGO_EMAIL_URL}" alt="MAPU" style="height:80px;display:block;margin:0 auto">
  </div>

  <!-- BODY -->
  <div style="padding:36px">
    <p style="font-size:17px;color:#1a1a1a;margin:0 0 12px">{t['email_greeting'].format(client=client)}</p>
    <p style="font-size:14px;color:#555;line-height:1.8;margin:0 0 24px">{t['email_intro']}</p>

    <p style="font-size:11px;color:#888;letter-spacing:2px;text-transform:uppercase;margin:0 0 12px">{t['email_details']}</p>
    <div style="background:#f8f8f8;padding:20px;margin:0 0 24px">
      <table style="width:100%;border-collapse:collapse">
        <tr><td style="padding:7px 0;font-size:13px;border-bottom:1px solid #eee;color:#888;width:45%">{t['email_checkin']}</td><td style="padding:7px 0;font-size:13px;border-bottom:1px solid #eee;color:#1a1a1a;font-weight:bold">{date_fmt(data['checkin'], lang)}</td></tr>
        <tr><td style="padding:7px 0;font-size:13px;border-bottom:1px solid #eee;color:#888">{t['email_checkout']}</td><td style="padding:7px 0;font-size:13px;border-bottom:1px solid #eee;color:#1a1a1a;font-weight:bold">{date_fmt(data['checkout'], lang)}</td></tr>
        <tr><td style="padding:7px 0;font-size:13px;border-bottom:1px solid #eee;color:#888">{t['email_nights']}</td><td style="padding:7px 0;font-size:13px;border-bottom:1px solid #eee;color:#1a1a1a;font-weight:bold">{nights}</td></tr>
        <tr><td style="padding:7px 0;font-size:13px;border-bottom:1px solid #eee;color:#888">{t['email_cabin']}</td><td style="padding:7px 0;font-size:13px;border-bottom:1px solid #eee;color:#1a1a1a;font-weight:bold">{cabins_str}</td></tr>
        <tr><td style="padding:7px 0;font-size:13px;color:#888">{t['email_guests']}</td><td style="padding:7px 0;font-size:13px;color:#1a1a1a;font-weight:bold">{guests_str}</td></tr>
      </table>
    </div>

    <table style="width:100%;background:#1a1a1a;margin:0 0 24px">
      <tr>
        <td style="padding:22px 24px;color:#fff;font-size:15px;font-weight:600;line-height:1.6;text-align:center">{t['email_open_pdf']}</td>
      </tr>
    </table>

    <p style="font-size:13px;color:#555;line-height:1.8;margin:0 0 8px">{t['email_cta']}</p>
    <p style="font-size:12px;color:#aaa;margin:0 0 24px">{t['email_validity']}</p>
    <hr style="border:none;border-top:1px solid #eee;margin:24px 0">
    <p style="font-size:13px;color:#555;line-height:1.8;margin:0">{t['email_signature']}<br>
    <a href="mailto:hola@mapuchile.com" style="color:#888">hola@mapuchile.com</a> · +569 58642354</p>
  </div>

  <!-- FOOTER -->
  <div style="background:#1a1a1a;padding:20px 36px;text-align:center">
    <a href="https://www.instagram.com/mapu_chile/" style="color:#ffffff;font-size:12px;text-decoration:none;letter-spacing:1px">@mapu_chile</a>
  </div>

</div>
</body>
</html>"""


MAKE_STORE_URL = "https://hook.us2.make.com/i05g74qnqjz2g3c48i3lm9n5skncvw3v"

def send_proposal_email(data, calcs, pdf_path, slug=None):
    import urllib.request

    email_cliente = data.get("email_cliente", "").strip()
    if not email_cliente:
        print("  ⚠ Email do cliente não informado — proposta não enviada para aprovação.")
        return False

    lang = data.get("lang", "pt")
    t    = TRANSLATIONS.get(lang, TRANSLATIONS["pt"])

    subject  = t["email_subject"].format(client=data["client_raw"])
    html     = _build_email_html(data, calcs, t)

    pdf_b64      = ""
    pdf_filename = ""
    if pdf_path and Path(pdf_path).exists():
        import base64
        with open(pdf_path, "rb") as f:
            pdf_b64 = base64.b64encode(f.read()).decode("utf-8")
        pdf_filename = Path(pdf_path).name

    import urllib.parse
    row_id       = slug or data.get("client", "proposta")
    checkin_str  = data["checkin"].strftime("%d/%m/%Y") if hasattr(data["checkin"], "strftime") else str(data["checkin"])
    checkout_str = data["checkout"].strftime("%d/%m/%Y") if hasattr(data["checkout"], "strftime") else str(data["checkout"])
    total_str    = fmt_clp(calcs.get("total_cc", 0))

    folder_url = ""
    if pdf_path:
        client_path = Path(pdf_path).parent.parent
        relative = str(client_path).replace(str(Path.home() / "Dropbox") + "/", "")
        folder_url = "https://www.dropbox.com/home/" + "/".join(
            urllib.parse.quote(p, safe="") for p in relative.split("/")
        )

    payload = json.dumps({
        "row_id":       row_id,
        "to":           email_cliente,
        "subject":      subject,
        "html":         html,
        "pdf_filename": pdf_filename,
        "pdf_base64":   pdf_b64,
        "client":       data.get("client_raw", data.get("client", "")),
        "checkin":      checkin_str,
        "checkout":     checkout_str,
        "total":        total_str,
        "folder_url":   folder_url,
    }).encode("utf-8")

    try:
        req = urllib.request.Request(
            MAKE_STORE_URL,
            data=payload,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=15) as resp:
            resp.read()
        print(f"  ✓ Proposta enviada para aprovação — aguardando decisão em hola@mapuchile.com")
        return True
    except Exception as e:
        print(f"  ✗ Erro ao enviar para aprovação: {e}")
        return False


# ── REMOVE LINKS EXTERNOS DO XLSX (evita diálogo de segurança do Excel) ────────
def _resolve_external_links(xlsx_path):
    """Substitui referências a links externos pelos valores atuais do arquivo fonte.
    Lê o PRECOS_CABINS.xlsx para obter preços atualizados; usa cache como fallback."""
    import zipfile, re, os, urllib.parse

    with zipfile.ZipFile(str(xlsx_path), 'r') as z:
        file_data = {name: z.read(name) for name in z.namelist()}

    ext_key = 'xl/externalLinks/externalLink1.xml'
    if ext_key not in file_data:
        return  # sem links externos, nada a fazer

    # Esta função só sabe resolver o link antigo (formato [1]PRECOS!$B$XX, sem
    # aspas/espaço/emoji no nome da aba). Link novo (ex: '[1]⚙️ INPUTS'!$B$6,
    # com aspas por causa do espaço/emoji) não bate com esse padrão — nesse
    # caso não mexer em nada e deixar o Excel resolver sozinho no osascript
    # (o dialog "Update Links" já é tratado lá).
    old_style_link = any(
        re.search(r'\[1\][A-Z_]+!\$', file_data[name].decode('utf-8', errors='ignore'))
        for name in file_data if name.startswith('xl/worksheets/')
    )
    if not old_style_link:
        return

    # Lê valores em cache do externalLink XML como fallback
    cached = {}
    for m in re.finditer(r'<cell r="([^"]+)"[^>]*><v>([^<]+)</v>',
                         file_data[ext_key].decode('utf-8')):
        cached[m.group(1)] = m.group(2)

    # Tenta ler preços atuais do arquivo fonte (PRECOS_CABINS.xlsx)
    ext_rels_key = 'xl/externalLinks/_rels/externalLink1.xml.rels'
    live_values = {}
    if ext_rels_key in file_data:
        rels_content = file_data[ext_rels_key].decode('utf-8')
        path_match = re.search(r'Target="([^"]+\.xlsx)"', rels_content)
        if path_match:
            src_path = urllib.parse.unquote(path_match.group(1))
            if os.path.exists(src_path):
                try:
                    src_wb = openpyxl.load_workbook(src_path, data_only=True)
                    # Descobre o nome da aba referenciada (ex: PRECOS)
                    sheet_match = re.search(r'\[1\]([A-Z_]+)!\$',
                                            file_data.get('xl/worksheets/sheet2.xml', b'').decode('utf-8', errors='ignore'))
                    sheet_name = sheet_match.group(1) if sheet_match else None
                    if sheet_name and sheet_name in src_wb.sheetnames:
                        ws = src_wb[sheet_name]
                        for cell_ref in cached:  # percorre só as células que o template usa
                            val = ws[cell_ref].value
                            if val is not None:
                                live_values[cell_ref] = str(val)
                        print(f"  ✓ Preços atualizados de {os.path.basename(src_path)}: {live_values}")
                except Exception as e:
                    print(f"  ⚠ Não foi possível ler preços de {src_path}: {e} — usando cache")

    values = {**cached, **live_values}  # live sobrescreve cache quando disponível

    # Substitui [1]PRECOS!$B$XX pelos valores nas planilhas
    for name in list(file_data.keys()):
        if not name.startswith('xl/worksheets/'):
            continue
        content = file_data[name].decode('utf-8', errors='ignore')
        if '[1]' not in content:
            continue
        content = re.sub(
            r'\[1\][A-Z_]+!\$([A-Z]+)\$(\d+)',
            lambda m: values.get(m.group(1) + m.group(2), m.group(0)),
            content
        )
        file_data[name] = content.encode('utf-8')

    # Remove arquivos de link externo
    for key in list(file_data.keys()):
        if 'externalLink' in key:
            del file_data[key]

    # Remove entradas em [Content_Types].xml
    ct_key = '[Content_Types].xml'
    if ct_key in file_data:
        ct = file_data[ct_key].decode('utf-8')
        ct = re.sub(r'<Override[^>]*externalLink[^>]*/>', '', ct)
        file_data[ct_key] = ct.encode('utf-8')

    # Remove relacionamento em xl/_rels/workbook.xml.rels
    rels_key = 'xl/_rels/workbook.xml.rels'
    if rels_key in file_data:
        rels = file_data[rels_key].decode('utf-8')
        rels = re.sub(r'<Relationship[^>]*externalLink[^>]*/>', '', rels)
        file_data[rels_key] = rels.encode('utf-8')

    # Remove <externalReferences> de xl/workbook.xml
    wb_key = 'xl/workbook.xml'
    if wb_key in file_data:
        wb = file_data[wb_key].decode('utf-8')
        wb = re.sub(r'<externalReferences>.*?</externalReferences>', '', wb, flags=re.DOTALL)
        file_data[wb_key] = wb.encode('utf-8')

    tmp = str(xlsx_path) + '.tmp'
    with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as z:
        for name, data in file_data.items():
            z.writestr(name, data)
    os.replace(tmp, str(xlsx_path))


# ── LER RESULTADOS DO EXCEL ────────────────────────────────────────────────────
def read_excel_results(xlsx_path, agency=False, data=None):
    import subprocess, time

    # Copia para pasta raiz (Excel já tem permissão), evita dialog "Grant Access"
    tmp_path = ORCAMENTOS / "_temp_calc.xlsx"
    shutil.copy2(str(xlsx_path), str(tmp_path))

    _resolve_external_links(tmp_path)

    # Remove lock file residual de sessões anteriores
    lock = ORCAMENTOS / "~$_temp_calc.xlsx"
    if lock.exists():
        try:
            lock.unlink()
        except Exception:
            pass

    script = f'''
tell application "Microsoft Excel"
    set fName to "{tmp_path}"
    try
        set wb to workbook (POSIX file fName as text)
        close wb saving no
    end try
    activate
    set wb to open workbook workbook file name fName
end tell

-- Dismiss any dialog immediately after open (key code 36 = Return = default button)
tell application "System Events"
    tell process "Microsoft Excel"
        repeat 8 times
            delay 1
            try
                repeat with w in (every window)
                    try
                        if exists (button "Keep Current Format" of w) then
                            click button "Keep Current Format" of w
                        end if
                    end try
                    try
                        if exists (button "Don't Update" of w) then
                            click button "Don't Update" of w
                        end if
                    end try
                    try
                        if exists (button "Update" of w) then
                            click button "Update" of w
                        end if
                    end try
                    try
                        if exists (button "Yes" of w) then
                            click button "Yes" of w
                        end if
                    end try
                end repeat
            end try
            -- Fallback: press Return to dismiss any remaining dialog
            try
                key code 36
            end try
        end repeat
    end tell
end tell

tell application "Microsoft Excel"
    delay 2
end tell

tell application "Microsoft Excel"
    delay 5
    save wb
    close wb saving yes
end tell
'''
    proc = subprocess.Popen(["osascript", "-e", script],
                            stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    stdout, stderr = proc.communicate(timeout=120)
    if proc.returncode != 0:
        raise RuntimeError(f"osascript falhou: {stderr.decode().strip()}")

    time.sleep(1)  # garante que o arquivo foi liberado antes de ler

    # Copia resultado de volta e limpa temp
    shutil.copy2(str(tmp_path), str(xlsx_path))
    try:
        tmp_path.unlink()
    except Exception:
        pass

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb["RESULTADO"]

    def v(cell):
        val = ws[cell].value
        return round(val) if val else 0

    lodging         = v("E11")
    food            = v("E13")
    beverages       = v("E15")   # E15 = bebidas (E14 era vazio)
    activities      = v("E17")
    transport       = v("E19")
    experiences     = v("E21")
    extras          = v("E23")   # E23 = extras (coordenação, limpeza etc.)
    quincho         = v("E25")   # E25 = quincho
    subtotal        = v("E27")
    total_cc        = v("E28")
    total_agency_cc = v("E31")
    agency_fee      = v("E33") if agency else 0

    total_final = total_agency_cc if agency else total_cc

    return {
        "lodging":          lodging,
        "food":             food,
        "beverages":        beverages,
        "activities":       activities,
        "transport":        transport,
        "experiences":      experiences,
        "extras":           extras,
        "quincho":          quincho,
        "mapu_team":        0,
        "agency_fee":       agency_fee,
        "iva":              0,
        "total":            subtotal,
        "total_bruto":      subtotal,
        "total_cc":         total_final,
        "total_agency_cc":  total_agency_cc,
    }


# ── PERSISTÊNCIA DE DATA ───────────────────────────────────────────────────────
def _save_data_json(data, client_path, slug):
    import json
    from datetime import date as date_type
    def _serial(obj):
        if isinstance(obj, date_type):
            return obj.isoformat()
        return str(obj)
    json_path = client_path / "1. ORCAMENTO" / f"{slug}_data.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=_serial)
    return json_path


def _load_data_json(json_path):
    import json
    from datetime import date as date_type
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)
    for key in ("checkin", "checkout"):
        if isinstance(data.get(key), str):
            data[key] = date_type.fromisoformat(data[key])
    # Restaura chaves inteiras em extras_sel
    for cat, items in data.get("extras_sel", {}).items():
        data["extras_sel"][cat] = {int(k): v for k, v in items.items()}
    return data


GUSTAVO_EMAIL = "hola@mapuchile.com"


def send_approval_notification(data, calcs, pdf_path, app_url="http://localhost:8502"):
    """Envia email para Gustavo com PDF + link de aprovação."""
    cfg = load_smtp_config()
    if not cfg:
        print("  ⚠ SMTP não configurado — notificação não enviada")
        return

    import urllib.parse
    slug    = data["checkin"].strftime("%Y%m%d") + "_" + data.get("client", "orcamento")
    checkin = data["checkin"].strftime("%d/%m/%Y")
    ag_name = data.get("agency_name", "Agência")
    total   = f"CLP {calcs['total_cc']:,.0f}".replace(",", ".")
    params = {
        "approve": slug,
        "cl":  data.get("client_raw", ""),
        "ci":  data["checkin"].strftime("%Y-%m-%d"),
        "co":  data["checkout"].strftime("%Y-%m-%d"),
        "n":   data.get("nights", 0),
        "cab": ",".join(data.get("cabins", {}).keys()),
        "pax": ",".join(
            f"{k}:{v.get('adults',0)}:{v.get('infants',0)}"
            for k, v in data.get("cabins", {}).items()
        ),
        "tot": int(calcs.get("total_cc", 0)),
        "ag":  data.get("agency_name", ""),
        "agc": data.get("agency_contact", ""),
        "age": data.get("agency_email", ""),
        "agp": data.get("agency_phone", ""),
        "agu": data.get("agency_user", ""),
        "cle": data.get("email_cliente", ""),
    }
    approve_url = f"{app_url}?{urllib.parse.urlencode(params)}"

    checkout   = data["checkout"].strftime("%d/%m/%Y")
    cabins_str = ", ".join(data["cabins"].keys())
    ag_contact = data.get("agency_contact", "")
    ag_email   = data.get("agency_email", "")
    ag_phone   = data.get("agency_phone", "")

    msg = MIMEMultipart()
    msg["From"]    = cfg["smtp_user"]
    msg["To"]      = GUSTAVO_EMAIL
    msg["Subject"] = f"[MAPU] Aprovação — {data['client_raw']} · {ag_name}"

    body = f"""<!DOCTYPE html><html><body style="margin:0;padding:0;background:#f4f4f4;font-family:Arial,sans-serif">
<div style="max-width:560px;margin:32px auto;background:#fff;border-radius:8px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,.08)">

  <!-- Header -->
  <div style="background:#1a1a1a;padding:24px 32px">
    <p style="margin:0;color:#fff;font-size:20px;font-weight:700;letter-spacing:1px">MAPU <span style="font-weight:300;font-style:italic">experiences lodge</span></p>
    <p style="margin:4px 0 0;color:#aaa;font-size:12px;letter-spacing:2px">NOVO ORÇAMENTO PARA APROVAÇÃO</p>
  </div>

  <!-- Body -->
  <div style="padding:28px 32px">

    <table style="width:100%;border-collapse:collapse;margin-bottom:24px">
      <tr><td style="padding:8px 0;border-bottom:1px solid #f0f0f0;color:#888;font-size:12px;text-transform:uppercase;letter-spacing:1px;width:120px">Cliente</td>
          <td style="padding:8px 0;border-bottom:1px solid #f0f0f0;font-size:15px;font-weight:600;color:#1a1a1a">{data['client_raw']}</td></tr>
      <tr><td style="padding:8px 0;border-bottom:1px solid #f0f0f0;color:#888;font-size:12px;text-transform:uppercase;letter-spacing:1px">Check-in</td>
          <td style="padding:8px 0;border-bottom:1px solid #f0f0f0;font-size:14px;color:#333">{checkin}</td></tr>
      <tr><td style="padding:8px 0;border-bottom:1px solid #f0f0f0;color:#888;font-size:12px;text-transform:uppercase;letter-spacing:1px">Check-out</td>
          <td style="padding:8px 0;border-bottom:1px solid #f0f0f0;font-size:14px;color:#333">{checkout} &nbsp;·&nbsp; {data['nights']} noite(s)</td></tr>
      <tr><td style="padding:8px 0;border-bottom:1px solid #f0f0f0;color:#888;font-size:12px;text-transform:uppercase;letter-spacing:1px">Cabanas</td>
          <td style="padding:8px 0;border-bottom:1px solid #f0f0f0;font-size:14px;color:#333">{cabins_str}</td></tr>
      <tr><td style="padding:8px 0;color:#888;font-size:12px;text-transform:uppercase;letter-spacing:1px">Total</td>
          <td style="padding:8px 0;font-size:18px;font-weight:700;color:#1a1a1a">{total}</td></tr>
    </table>

    <p style="margin:0 0 6px;color:#888;font-size:12px;text-transform:uppercase;letter-spacing:1px">Solicitado por</p>
    <p style="margin:0 0 24px;font-size:14px;color:#444">{ag_name} &nbsp;·&nbsp; {ag_contact}<br>
    <a href="mailto:{ag_email}" style="color:#888">{ag_email}</a> &nbsp;·&nbsp; {ag_phone}</p>

    <a href="{approve_url}" style="display:inline-block;background:#1a1a1a;color:#fff;padding:14px 28px;text-decoration:none;border-radius:4px;font-size:14px;font-weight:600;letter-spacing:1px">
      ✓ &nbsp;APROVAR E ENVIAR PDF
    </a>

    <p style="margin:20px 0 0;font-size:12px;color:#aaa">PDF completo em anexo para revisão antes de aprovar.</p>
  </div>

  <!-- Footer -->
  <div style="background:#f8f8f8;padding:16px 32px;border-top:1px solid #eee">
    <p style="margin:0;font-size:11px;color:#bbb;text-align:center">MAPU Experiences Lodge &nbsp;·&nbsp; hola@mapuchile.com &nbsp;·&nbsp; +569 58642354</p>
  </div>

</div>
</body></html>"""
    msg.attach(MIMEText(body, "html"))

    if pdf_path and Path(pdf_path).exists():
        with open(pdf_path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename={Path(pdf_path).name}")
        msg.attach(part)

    try:
        with smtplib.SMTP(cfg["smtp_host"], cfg["smtp_port"]) as server:
            server.starttls()
            server.login(cfg["smtp_user"], cfg["smtp_pass"])
            server.sendmail(cfg["smtp_user"], GUSTAVO_EMAIL, msg.as_string())
        print(f"  ✓ Notificação enviada para {GUSTAVO_EMAIL}")
    except Exception as e:
        print(f"  ⚠ Erro ao enviar notificação: {e}")


def send_approved_budget_email(data, calcs, pdf_path):
    """Envia PDF aprovado — pro email do cliente se informado, senão pro contato da agência."""
    cfg = load_smtp_config()
    if not cfg:
        return False

    client_email = data.get("email_cliente", "")
    ag_email     = data.get("agency_email", "")
    ag_contact   = data.get("agency_contact", "")
    ag_name      = data.get("agency_name", "")
    to_email     = client_email or ag_email
    if not to_email:
        print("  ⚠ Nenhum email de destino encontrado (cliente ou agência)")
        return False
    greeting_name = data.get("client_raw", "") if client_email else ag_contact

    def _fmt(d):
        if hasattr(d, "strftime"):
            return d.strftime("%d/%m/%Y")
        # string ISO "YYYY-MM-DD"
        try:
            from datetime import datetime
            return datetime.strptime(str(d), "%Y-%m-%d").strftime("%d/%m/%Y")
        except Exception:
            return str(d)
    checkin  = _fmt(data.get("checkin", ""))
    checkout = _fmt(data.get("checkout", ""))
    total    = f"CLP {calcs['total_cc']:,.0f}".replace(",", ".")

    msg = MIMEMultipart()
    msg["From"]    = cfg["smtp_user"]
    msg["To"]      = to_email
    msg["Subject"] = f"MAPU — Orçamento aprovado: {data['client_raw']}"

    nights = data.get("nights", "—")
    # Linhas por cabana com pax
    cabin_rows = ""
    for cab, pax in data.get("cabins", {}).items():
        adults  = pax.get("adults", 0)  if isinstance(pax, dict) else 0
        infants = pax.get("infants", 0) if isinstance(pax, dict) else 0
        if adults or infants:
            pax_label = f"{adults} adulto(s)" + (f" + {infants} criança(s)" if infants else "")
        else:
            pax_label = "—"
        cabin_rows += f"""
      <tr>
        <td style="padding:8px 0;border-bottom:1px solid #f0f0f0;color:#888;font-size:12px;text-transform:uppercase;letter-spacing:1px;width:120px">{cab}</td>
        <td style="padding:8px 0;border-bottom:1px solid #f0f0f0;font-size:13px;color:#333">{pax_label}</td>
      </tr>"""

    body = f"""<!DOCTYPE html><html><body style="margin:0;padding:0;background:#f4f4f4;font-family:Arial,sans-serif">
<div style="max-width:560px;margin:32px auto;background:#fff;border-radius:8px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,.08)">

  <!-- Header -->
  <div style="background:#1a1a1a;padding:24px 32px">
    <p style="margin:0;color:#fff;font-size:20px;font-weight:700;letter-spacing:1px">MAPU <span style="font-weight:300;font-style:italic">experiences lodge</span></p>
    <p style="margin:4px 0 0;color:#aaa;font-size:12px;letter-spacing:2px">ORÇAMENTO APROVADO</p>
  </div>

  <!-- Body -->
  <div style="padding:28px 32px">
    <p style="margin:0 0 24px;font-size:14px;color:#555;line-height:1.6">
      {"Olá <b>" + greeting_name + "</b>, segue em anexo seu orçamento MAPU aprovado." if client_email else "Olá <b>" + greeting_name + "</b>, o orçamento para <b>" + data['client_raw'] + "</b> foi aprovado.<br>Segue em anexo o PDF para envio ao cliente."}
    </p>

    <table style="width:100%;border-collapse:collapse;margin-bottom:24px">
      <tr>
        <td style="padding:8px 0;border-bottom:1px solid #f0f0f0;color:#888;font-size:12px;text-transform:uppercase;letter-spacing:1px;width:120px">Cliente</td>
        <td style="padding:8px 0;border-bottom:1px solid #f0f0f0;font-size:15px;font-weight:600;color:#1a1a1a">{data['client_raw']}</td>
      </tr>
      <tr>
        <td style="padding:8px 0;border-bottom:1px solid #f0f0f0;color:#888;font-size:12px;text-transform:uppercase;letter-spacing:1px">Check-in</td>
        <td style="padding:8px 0;border-bottom:1px solid #f0f0f0;font-size:14px;color:#333">{checkin}</td>
      </tr>
      <tr>
        <td style="padding:8px 0;border-bottom:1px solid #f0f0f0;color:#888;font-size:12px;text-transform:uppercase;letter-spacing:1px">Check-out</td>
        <td style="padding:8px 0;border-bottom:1px solid #f0f0f0;font-size:14px;color:#333">{checkout} &nbsp;·&nbsp; {nights} noite(s)</td>
      </tr>
      {cabin_rows}
      <tr>
        <td style="padding:12px 0 4px;color:#888;font-size:12px;text-transform:uppercase;letter-spacing:1px">Total (c/ CC 4%)</td>
        <td style="padding:12px 0 4px;font-size:18px;font-weight:700;color:#1a1a1a">{total}</td>
      </tr>
    </table>
  </div>

  <!-- Footer -->
  <div style="background:#f8f8f8;padding:16px 32px;border-top:1px solid #eee">
    <p style="margin:0;font-size:11px;color:#bbb;text-align:center">
      MAPU Experiences Lodge &nbsp;·&nbsp; hola@mapuchile.com &nbsp;·&nbsp; +569 58642354
    </p>
  </div>

</div>
</body></html>"""
    msg.attach(MIMEText(body, "html"))

    if pdf_path and Path(pdf_path).exists():
        with open(pdf_path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename={Path(pdf_path).name}")
        msg.attach(part)

    try:
        with smtplib.SMTP(cfg["smtp_host"], cfg["smtp_port"]) as server:
            server.starttls()
            server.login(cfg["smtp_user"], cfg["smtp_pass"])
            server.sendmail(cfg["smtp_user"], to_email, msg.as_string())
        print(f"  ✓ PDF aprovado enviado para {to_email}")
        return True
    except Exception as e:
        print(f"  ⚠ Erro ao enviar: {e}")
        return False


def download_pdf_from_dropbox(slug, agency_user):
    """Baixa só o PDF do Dropbox para anexar no email de aprovação."""
    import tempfile
    dbx = _get_dropbox_client()
    if not dbx:
        return None
    try:
        year, month = slug[:4], slug[4:6]
        ag = agency_user.upper()
        # tenta os dois cases do nome da subpasta
        entries = None
        for sub in (f"{DROPBOX_ROOT}/{year}/{month}/{slug}_{ag}/1. orcamento",
                    f"{DROPBOX_ROOT}/{year}/{month}/{slug}_{ag}/1. ORCAMENTO"):
            try:
                entries = dbx.files_list_folder(sub).entries
                break
            except Exception:
                continue
        if entries is None:
            print(f"  ⚠ Subpasta 1.ORCAMENTO não encontrada para {slug}_{ag}")
            return None
        pdf_entry = next((e for e in entries if e.name.endswith(".pdf")), None)
        if not pdf_entry:
            return None
        tmp_dir = Path(tempfile.mkdtemp())
        pdf_path = tmp_dir / pdf_entry.name
        # obtém link temporário e baixa via urllib (mais robusto em nuvem)
        import urllib.request
        link = dbx.files_get_temporary_link(pdf_entry.path_lower).link
        urllib.request.urlretrieve(link, str(pdf_path))
        return pdf_path if pdf_path.exists() else None
    except Exception as e:
        print(f"  ⚠ Erro ao baixar PDF do Dropbox: {e}")
        return None


def find_budget_files(slug):
    """Encontra PDF e data.json de um orçamento pelo slug."""
    if IS_CLOUD:
        return _find_budget_files_dropbox(slug)
    for json_path in ORCAMENTOS.rglob(f"{slug}_data.json"):
        pdf_candidates = list(json_path.parent.glob("*.pdf"))
        pdf_path = pdf_candidates[0] if pdf_candidates else None
        return json_path, pdf_path
    return None, None


def _find_budget_files_dropbox(slug):
    """Baixa data.json e PDF do Dropbox navegando pela estrutura de pastas."""
    import tempfile
    dbx = _get_dropbox_client()
    if not dbx:
        print("  ⚠ Dropbox client não disponível")
        return None, None
    try:
        # slug formato: YYYYMMDD_CLIENT — extrai ano e mês
        year  = slug[:4]
        month = slug[4:6]
        parent = f"/{year}/{month}"

        # Lista subpastas do mês para achar a que começa com o slug
        result = dbx.files_list_folder(parent)
        target = None
        for entry in result.entries:
            if entry.name.startswith(slug):
                target = entry.path_lower
                break

        if not target:
            print(f"  ⚠ Pasta {slug} não encontrada em Dropbox:{parent}")
            return None, None

        # Lista a subpasta "1. orcamento"
        orcamento = f"{target}/1. orcamento"
        try:
            entries = dbx.files_list_folder(orcamento).entries
        except Exception:
            # tenta com case original
            orcamento = f"{target}/1. ORCAMENTO"
            entries = dbx.files_list_folder(orcamento).entries

        tmp_dir = Path(tempfile.mkdtemp())

        json_entry = next((e for e in entries if e.name.endswith("_data.json")), None)
        pdf_entry  = next((e for e in entries if e.name.endswith(".pdf")), None)

        json_path = None
        if json_entry:
            json_path = tmp_dir / json_entry.name
            _, resp = dbx.files_download(json_entry.path_lower)
            json_path.write_bytes(resp.content)

        pdf_path = None
        if pdf_entry:
            pdf_path = tmp_dir / pdf_entry.name
            _, resp = dbx.files_download(pdf_entry.path_lower)
            pdf_path.write_bytes(resp.content)

        return json_path, pdf_path

    except Exception as e:
        print(f"  ⚠ Dropbox find error: {e}")
        return None, None


def _list_recent_orcamentos(n=10):
    """Retorna lista de (xlsx_path, json_path) dos orçamentos mais recentes."""
    results = []
    for json_path in sorted(
        ORCAMENTOS.rglob("*_data.json"),
        key=lambda p: p.stat().st_mtime, reverse=True
    )[:n]:
        xlsx_path = json_path.parent / json_path.name.replace("_data.json", ".xlsx")
        if xlsx_path.exists():
            results.append((xlsx_path, json_path))
    return results


def regenerate_pdf():
    """Modo 2: lê Excel já ajustado + data.json e regera só o PDF."""
    print("\n── REGENERAR PDF ────────────────────────────────────")
    recentes = _list_recent_orcamentos()
    if not recentes:
        print("  Nenhum orçamento encontrado com data.json.")
        return

    for i, (xlsx, _) in enumerate(recentes, 1):
        slug = xlsx.stem
        mtime = xlsx.stat().st_mtime
        from datetime import datetime
        dt = datetime.fromtimestamp(mtime).strftime("%d/%m %H:%M")
        print(f"  {i:>2}. {slug}  ({dt})")

    sel = input(f"\n  Selecionar (1-{len(recentes)}): ").strip()
    try:
        idx = int(sel) - 1
        xlsx_path, json_path = recentes[idx]
    except (ValueError, IndexError):
        print("  Seleção inválida.")
        return

    print(f"\n  Carregando: {xlsx_path.name}")
    data = _load_data_json(json_path)

    prices = load_precos(checkin_year=data["checkin"].year, checkin_month=data["checkin"].month)
    calcs  = calculate(data, prices)

    print("  Lendo totais do Excel...")
    try:
        excel_vals = read_excel_results(xlsx_path, agency=data.get("agency", False), data=data)
        calcs.update(excel_vals)
        calcs["usd_ref"]   = calcs["total_cc"] / data.get("exchange", 900)
        calcs["per_adult"] = calcs["total_cc"] / max(data.get("total_adults", 1), 1)
    except Exception as e:
        print(f"  ⚠ Não foi possível ler Excel ({e}) — usando estimativa Python.")

    client_path = xlsx_path.parent.parent
    slug        = xlsx_path.stem
    prop_num    = next_prop_num(client_path)
    pdf_path    = generate_pdf(data, calcs, client_path, slug, prop_num)
    print(f"\n  ✓ PDF gerado: {pdf_path.name}")

    if data.get("email_cliente"):
        send = input("\n  Enviar proposta para aprovação? (s/n) [s]: ").strip().lower()
        if send != "n":
            send_proposal_email(data, calcs, pdf_path, slug=slug)


# ── MAIN ───────────────────────────────────────────────────────────────────────
def next_prop_num(client_path):
    orcamento_dir = client_path / "1. ORCAMENTO"
    if not orcamento_dir.exists():
        return "01"
    existing = list(orcamento_dir.glob("Prop*_*.pdf"))
    return f"{len(existing) + 1:02d}"


def _get_dropbox_client():
    """Retorna cliente Dropbox autenticado via refresh token."""
    import dropbox, json as _json
    cfg_path = Path(__file__).parent / "dropbox_config.json"
    if cfg_path.exists():
        cfg = _json.loads(cfg_path.read_text())
    else:
        try:
            import streamlit as st
            s = st.secrets["dropbox"]
            cfg = {"app_key": s["app_key"], "app_secret": s["app_secret"],
                   "refresh_token": s["refresh_token"]}
        except Exception:
            return None
    return dropbox.Dropbox(
        oauth2_refresh_token=cfg["refresh_token"],
        app_key=cfg["app_key"],
        app_secret=cfg["app_secret"],
    )


def send_registration_request(agency_data, app_url):
    """Envia email para Gustavo pedindo aprovação de nova agência."""
    import urllib.parse
    cfg = load_smtp_config()
    if not cfg:
        return False
    params = {
        "register": "1",
        "ag_name":    agency_data.get("agency_name", ""),
        "ag_contact": agency_data.get("contact", ""),
        "ag_email":   agency_data.get("email", ""),
        "ag_phone":   agency_data.get("phone", ""),
    }
    approve_url = f"{app_url}?{urllib.parse.urlencode(params)}"
    msg = MIMEMultipart()
    msg["From"]    = cfg["smtp_user"]
    msg["To"]      = GUSTAVO_EMAIL
    msg["Subject"] = "NOVO USUÁRIO | ACESSO APP ORÇAMENTOS MAPU LODGE"
    body = f"""<!DOCTYPE html><html><body style="font-family:Arial,sans-serif;background:#f4f4f4;padding:32px">
<div style="max-width:520px;margin:0 auto;background:#fff;border-radius:8px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,.08)">
  <div style="background:#1a1a1a;padding:20px 28px">
    <p style="margin:0;color:#fff;font-size:18px;font-weight:700">MAPU <span style="font-weight:300;font-style:italic">experiences lodge</span></p>
    <p style="margin:4px 0 0;color:#aaa;font-size:11px;letter-spacing:2px">NOVA SOLICITAÇÃO DE ACESSO</p>
  </div>
  <div style="padding:24px 28px">
    <table style="width:100%;border-collapse:collapse">
      <tr><td style="padding:6px 0;color:#888;font-size:12px;width:100px">Agência</td><td style="padding:6px 0;font-weight:600">{agency_data.get('agency_name','')}</td></tr>
      <tr><td style="padding:6px 0;color:#888;font-size:12px">Contato</td><td style="padding:6px 0">{agency_data.get('contact','')}</td></tr>
      <tr><td style="padding:6px 0;color:#888;font-size:12px">Email</td><td style="padding:6px 0">{agency_data.get('email','')}</td></tr>
      <tr><td style="padding:6px 0;color:#888;font-size:12px">Telefone</td><td style="padding:6px 0">{agency_data.get('phone','')}</td></tr>
    </table>
    <div style="margin-top:24px">
      <a href="{approve_url}" style="display:inline-block;background:#1a1a1a;color:#fff;padding:12px 24px;text-decoration:none;border-radius:4px;font-size:13px;font-weight:600">✓ Aprovar e Criar Acesso</a>
    </div>
  </div>
</div></body></html>"""
    msg.attach(MIMEText(body, "html"))
    try:
        with smtplib.SMTP(cfg["smtp_host"], cfg["smtp_port"]) as server:
            server.starttls()
            server.login(cfg["smtp_user"], cfg["smtp_pass"])
            server.sendmail(cfg["smtp_user"], GUSTAVO_EMAIL, msg.as_string())
        return True
    except Exception as e:
        print(f"  ⚠ Erro ao enviar registro: {e}")
        return False


def send_agency_credentials(agency_data, username, password):
    """Envia email para a agência com login e senha criados."""
    cfg = load_smtp_config()
    if not cfg:
        return False
    ag_email = agency_data.get("email", "")
    if not ag_email:
        return False
    msg = MIMEMultipart()
    msg["From"]    = cfg["smtp_user"]
    msg["To"]      = ag_email
    msg["Subject"] = "MAPU — Seu acesso foi aprovado"
    body = f"""<!DOCTYPE html><html><body style="font-family:Arial,sans-serif;background:#f4f4f4;padding:32px">
<div style="max-width:520px;margin:0 auto;background:#fff;border-radius:8px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,.08)">
  <div style="background:#1a1a1a;padding:20px 28px">
    <p style="margin:0;color:#fff;font-size:18px;font-weight:700">MAPU <span style="font-weight:300;font-style:italic">experiences lodge</span></p>
  </div>
  <div style="padding:24px 28px">
    <p>Olá <b>{agency_data.get('contact','')}</b>,</p>
    <p>Seu acesso ao sistema de orçamentos MAPU foi aprovado. Use as credenciais abaixo:</p>
    <table style="width:100%;border-collapse:collapse;margin:16px 0">
      <tr><td style="padding:8px;background:#f8f8f8;color:#888;font-size:12px;width:80px">Usuário</td><td style="padding:8px;background:#f8f8f8;font-weight:700;font-size:16px">{username}</td></tr>
      <tr><td style="padding:8px;color:#888;font-size:12px">Senha</td><td style="padding:8px;font-weight:700;font-size:16px">{password}</td></tr>
    </table>
    <a href="https://mapu-orcamentos.streamlit.app" style="display:inline-block;background:#1a1a1a;color:#fff;padding:12px 24px;text-decoration:none;border-radius:4px;font-size:13px;font-weight:600">Acessar o sistema</a>
    <p style="margin-top:20px;font-size:12px;color:#aaa">MAPU Experiences Lodge · hola@mapuchile.com · +569 58642354</p>
  </div>
</div></body></html>"""
    msg.attach(MIMEText(body, "html"))
    try:
        with smtplib.SMTP(cfg["smtp_host"], cfg["smtp_port"]) as server:
            server.starttls()
            server.login(cfg["smtp_user"], cfg["smtp_pass"])
            server.sendmail(cfg["smtp_user"], ag_email, msg.as_string())
        return True
    except Exception as e:
        print(f"  ⚠ Erro ao enviar credenciais: {e}")
        return False


def generate_agency_credentials(agency_name, existing_users):
    """Gera username único e senha aleatória."""
    import re, random, string
    base = re.sub(r"[^a-z0-9]", "", agency_name.lower())[:12] or "agencia"
    username = base
    i = 2
    while username in existing_users:
        username = f"{base}{i}"
        i += 1
    password = "".join(random.choices(string.ascii_letters + string.digits, k=10))
    return username, password


def _dropbox_access_token():
    """Obtém access token fresco via OAuth2 refresh (HTTP direto)."""
    import urllib.request, urllib.parse
    local = Path(__file__).parent / "dropbox_config.json"
    if local.exists():
        cfg = json.loads(local.read_text())
    else:
        try:
            import streamlit as st
            sec = st.secrets["dropbox"]
            cfg = {"app_key": sec["app_key"], "app_secret": sec["app_secret"],
                   "refresh_token": sec["refresh_token"]}
        except Exception:
            return None
    try:
        body = urllib.parse.urlencode({
            "grant_type": "refresh_token",
            "refresh_token": cfg["refresh_token"],
            "client_id": cfg["app_key"],
            "client_secret": cfg["app_secret"],
        }).encode()
        req = urllib.request.Request("https://api.dropboxapi.com/oauth2/token",
                                     data=body, method="POST")
        with urllib.request.urlopen(req, timeout=15) as r:
            return json.loads(r.read())["access_token"]
    except Exception as e:
        print(f"_dropbox_access_token erro: {e}")
        return None


def load_agencies_dropbox():
    """Carrega agencies.json do Dropbox via SDK."""
    dbx = _get_dropbox_client()
    if not dbx:
        print("load_agencies_dropbox: sem cliente")
        return {}
    try:
        _, response = dbx.files_download(f"{DROPBOX_ROOT}/config/agencies.json")
        data = json.loads(response.content.decode("utf-8"))
        print(f"load_agencies_dropbox: {len(data)} agência(s)")
        return data
    except Exception as e:
        print(f"load_agencies_dropbox erro: {e}")
        return {}


def save_agencies_dropbox(agencies):
    """Salva agencies.json em /config/agencies.json no Dropbox."""
    dbx = _get_dropbox_client()
    if not dbx:
        return False
    try:
        import dropbox as _dbx
        data = json.dumps(agencies, ensure_ascii=False, indent=2).encode("utf-8")
        dbx.files_upload(data, f"{DROPBOX_ROOT}/config/agencies.json",
                         mode=_dbx.files.WriteMode.overwrite)
        return True
    except Exception as e:
        print(f"  ⚠ Erro ao salvar agencies: {e}")
        return False


def upload_budget_to_dropbox(pdf_path, xlsx_path, data):
    """Faz upload do PDF, Excel e data.json do orçamento para o Dropbox.
    Retorna (True, "") em sucesso, (False, motivo) em falha."""
    import dropbox as _dbx
    dbx = _get_dropbox_client()
    if not dbx:
        msg = "dropbox_config.json / Secrets do Dropbox não encontrados — upload ignorado"
        print(f"  ⚠ {msg}")
        return False, msg

    year  = data["checkin"].strftime("%Y")
    month = data["checkin"].strftime("%m")
    slug  = data["checkin"].strftime("%Y%m%d") + "_" + data["client"]
    ag    = data.get("agency_user", "agencia").upper()
    client_folder = f"{DROPBOX_ROOT}/{year}/{month}/{slug}_{ag}"
    folder = f"{client_folder}/1. ORCAMENTO"

    # Cria as demais subpastas (vazias) — "1. ORCAMENTO" é criada implicitamente pelo upload dos arquivos
    for sub in SUBFOLDERS[1:]:
        try:
            dbx.files_create_folder_v2(f"{client_folder}/{sub}")
        except _dbx.exceptions.ApiError:
            pass  # já existe

    # Deriva data.json a partir do pdf_path (mesma pasta, mesmo slug)
    data_json = None
    if pdf_path:
        candidate = Path(pdf_path).parent / f"{slug}_data.json"
        if candidate.exists():
            data_json = candidate

    try:
        for local_path in (pdf_path, xlsx_path, data_json):
            if not local_path or not Path(local_path).exists():
                continue
            dbx_path = f"{folder}/{Path(local_path).name}"
            with open(local_path, "rb") as f:
                dbx.files_upload(f.read(), dbx_path,
                                 mode=_dbx.files.WriteMode.overwrite)
            print(f"  ✓ Dropbox: {dbx_path}")
    except Exception as e:
        msg = f"Falha no upload para o Dropbox: {e}"
        print(f"  ⚠ {msg}")
        return False, msg

    return True, ""


def main():
    print("\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    print("   MAPU — GERADOR DE ORÇAMENTO")
    print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    print("    1 = Novo orçamento")
    print("    2 = Regenerar PDF (Excel já ajustado)")
    modo = input("  Modo (default: 1): ").strip()
    if modo == "2":
        regenerate_pdf()
        return

    data   = collect_inputs()
    prices = load_precos(checkin_year=data["checkin"].year, checkin_month=data["checkin"].month)

    if not data["meals"].get("breakfast"):
        mp  = prices.get("meal_prices") or {}
        bkf = mp.get("breakfast", {"adult": 24_000, "child": 18_000})
        t   = TRANSLATIONS.get(data.get("lang", "pt"), TRANSLATIONS["pt"])
        data["food_notes"] = [
            t["food_note_breakfast"].format(
                adult=f"{bkf['adult']:,.0f}".replace(",", "."),
                child=f"{bkf['child']:,.0f}".replace(",", "."),
            ),
            t["food_note_dinner"],
        ]

    calcs  = calculate(data, prices)

    print("\n── PRÉ-VISUALIZAÇÃO ─────────────────────────────────")
    print(f"  Alojamento:         {fmt_clp(calcs['lodging'])}")
    print(f"  Alimentação:        {fmt_clp(calcs['food'])}")
    print(f"  Subtotal:           {fmt_clp(calcs['total'])}")
    if calcs.get("agency_fee", 0) > 0:
        print(f"  Taxa agência:       {fmt_clp(calcs['agency_fee'])}")
    print(f"  TOTAL (c/ 5% CC):   {fmt_clp(calcs['total_cc'])}")
    print(f"  IVA 19% (informativo):{fmt_clp(round(calcs['total'] * 0.19))}")
    print(f"  Ref USD:            {fmt_usd(calcs['usd_ref'])} USD")
    print(f"  Por adulto:         {fmt_clp(calcs['per_adult'])}")

    confirm = input("\n  Gerar orçamento? (s/n) [s]: ").strip().lower()
    if confirm == "n":
        print("Cancelado.\n")
        return

    print("\n── GERANDO ──────────────────────────────────────────")
    client_path, slug = create_folders(data)
    prop_num   = next_prop_num(client_path)
    xlsx_path  = populate_excel(data, client_path, slug, calcs)
    _save_data_json(data, client_path, slug)
    print("  Lendo totais do Excel...")
    try:
        excel_vals = read_excel_results(xlsx_path, agency=data.get("agency", False), data=data)
        calcs.update(excel_vals)
        calcs["usd_ref"]   = calcs["total_cc"] / data["exchange"]
        calcs["per_adult"] = calcs["total_cc"] / max(data["total_adults"], 1)
    except Exception as e:
        print(f"  ⚠ Não foi possível ler Excel ({e}) — usando estimativa Python.")
    pdf_path = generate_pdf(data, calcs, client_path, slug, prop_num)

    if data.get("email_cliente"):
        send = prompt("\n  Enviar proposta para aprovação? (s/n)", "s").lower()
        if send != "n":
            send_proposal_email(data, calcs, pdf_path, slug=slug)

    print(f"\n  Pasta: {client_path}")
    print("  Abra o Excel — as fórmulas calculam ao abrir.\n")


if __name__ == "__main__":
    main()
